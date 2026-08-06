import os
import math
import time
import uuid
import datetime
import threading
import random
import pandas as pd
in_memory_jobs = []

MAX_DISTANCE_LIMIT = 800.0
_distance_cache = {}
OUTPUT_FOLDER = 'outputs'
os.makedirs(OUTPUT_FOLDER, exist_ok=True)





def parse_excel_nodes(file_path_or_df):
    if isinstance(file_path_or_df, pd.DataFrame):
        df = file_path_or_df
    else:
        df = pd.read_excel(file_path_or_df)
    
    nodes = []
    
    # Clean string helper
    def clean_str(val, default=''):
        if pd.isna(val):
            return default
        return str(val).strip()
        
    # Clean float helper
    def clean_float(val, default=0.0):
        if pd.isna(val):
            return default
        try:
            return float(val)
        except ValueError:
            return default

    # Group by node_id preserving order of appearance
    grouped = df.groupby('node_id', sort=False)
    
    for node_id, group in grouped:
        if pd.isna(node_id) or not str(node_id).strip():
            continue
        
        node_id = str(node_id).strip()
        first_row = group.iloc[0]
        
        node_type = clean_str(first_row.get('type')).lower()
        
        # Resolve latitude and longitude columns dynamically and case-insensitively
        raw_lat = None
        raw_lng = None
        for col in first_row.index:
            col_lower = str(col).lower().strip()
            if col_lower in ('lat', 'latitude'):
                raw_lat = first_row[col]
            elif col_lower in ('lng', 'long', 'longitude'):
                raw_lng = first_row[col]
                
        node = {
            'id': node_id,
            'name': clean_str(first_row.get('name'), node_id),
            'type': node_type,
            'subtype': clean_str(first_row.get('subtype')),
            'lat': clean_float(raw_lat) if raw_lat is not None else clean_float(first_row.get('lat')),
            'lng': clean_float(raw_lng) if raw_lng is not None else clean_float(first_row.get('lng')),
            'network_id': clean_str(first_row.get('network_id'))
        }
        
        if node_type == 'hub':
            products_dict = {}
            for _, row in group.iterrows():
                commodity = clean_str(row.get('commodity'))
                capacity = clean_float(row.get('capacity'))
                cost = clean_float(row.get('cost'))
                if commodity and capacity > 0:
                    if commodity not in products_dict:
                        products_dict[commodity] = {'capacity': 0.0, 'processing_cost': cost}
                    products_dict[commodity]['capacity'] += capacity
                    products_dict[commodity]['processing_cost'] = cost
            node['products'] = [{
                'type': k,
                'capacity': v['capacity'],
                'processing_cost': v['processing_cost']
            } for k, v in products_dict.items()]
            
        elif node_type == 'plant':
            inflow_dict = {}
            yield_dict = {}
            demand_dict = {}
            for _, row in group.iterrows():
                commodity = clean_str(row.get('commodity'))
                capacity = clean_float(row.get('capacity'))
                cost = clean_float(row.get('cost'))
                yield_val = clean_float(row.get('yield'))
                demand = clean_float(row.get('demand'))
                price = clean_float(row.get('price'))
                
                if capacity > 0 and commodity:
                    if commodity not in inflow_dict:
                        inflow_dict[commodity] = {'capacity': 0.0, 'processing_cost': cost}
                    inflow_dict[commodity]['capacity'] += capacity
                    inflow_dict[commodity]['processing_cost'] = cost
                if yield_val > 0 and commodity:
                    yield_dict[commodity] = yield_dict.get(commodity, 0.0) + yield_val
                if demand > 0 and commodity:
                    if commodity not in demand_dict:
                        demand_dict[commodity] = {'demand': 0.0, 'price': price, 'processing_cost': cost}
                    demand_dict[commodity]['demand'] += demand
                    demand_dict[commodity]['price'] = price
                    demand_dict[commodity]['processing_cost'] = cost
            
            node['inflow_milks'] = [{
                'type': k,
                'capacity': v['capacity'],
                'processing_cost': v['processing_cost']
            } for k, v in inflow_dict.items()]
            node['products'] = [{'type': k, 'yield': v} for k, v in yield_dict.items()]
            node['demands'] = [{
                'type': k,
                'demand': v['demand'],
                'price': v['price'],
                'processing_cost': v['processing_cost']
            } for k, v in demand_dict.items()]
            node['capacity'] = sum(m['capacity'] for m in node['inflow_milks']) if node['inflow_milks'] else sum(d['demand'] for d in node['demands'])
            node['processing_cost'] = node['inflow_milks'][0]['processing_cost'] if node['inflow_milks'] else (node['demands'][0]['processing_cost'] if node['demands'] else 0.40)
            

        nodes.append(node)
    return nodes

# Helper to determine raw milk type from finished product type
def get_milk_type_for_product(product_type):
    ptype = (product_type or '').strip()
    if ptype.endswith(' Cheese'):
        return ptype[:-7] + ' Milk'
    ptype_lower = ptype.lower()
    if 'buffalo' in ptype_lower:
        return 'Buffalo Milk'
    return 'Cow Milk'

# Helper to serialize MongoDB object
def serialize_node(node):
    node_dict = dict(node)
    if '_id' in node_dict:
        node_dict['_id'] = str(node_dict['_id'])
    return node_dict



def get_optimal_vehicles(flow, vehicle_limits, caps=None, caps_ranges=None, **kwargs):
    if caps is None:
        caps = {'7 L': 7000.0, '10 L': 10000.0, '12L': 12000.0, '15 L': 15000.0, '18 L': 18000.0, 'V30': 35000.0, 'V35': 35000.0}
    if caps_ranges is None:
        caps_ranges = {k: (v * 0.8, v) for k, v in caps.items()}
        
    if isinstance(vehicle_limits, dict) and 'limits' in vehicle_limits:
        vehicle_limits = vehicle_limits['limits']

    if flow < 6000:
        return {}

    sorted_caps = sorted(caps.items(), key=lambda x: x[1]) # ascending
    result = {k: 0 for k in caps}
    remaining = flow
    
    while remaining >= 6000:
        # 1. Find if it fits exactly within a truck's range
        range_fit_truck = None
        for name, val in sorted_caps:
            limit = int(vehicle_limits.get(name, 1000000))
            if limit > result[name]:
                from_val, to_val = caps_ranges.get(name, (val * 0.8, val))
                if from_val <= remaining <= to_val:
                    range_fit_truck = name
                    break
                    
        if range_fit_truck:
            result[range_fit_truck] += 1
            remaining = 0
            break
            
        # 2. Otherwise, find the largest truck that fits completely within the remaining
        round_down_truck = None
        round_down_val = None
        for name, val in reversed(sorted_caps):
            limit = int(vehicle_limits.get(name, 1000000))
            if limit > result[name] and val <= remaining:
                round_down_truck = name
                round_down_val = val
                break
                
        if round_down_truck:
            result[round_down_truck] += 1
            remaining -= round_down_val
        else:
            # 3. Remaining is < smallest truck. But >= 6000. So dispatch smallest available truck.
            smallest_truck = None
            for name, val in sorted_caps:
                limit = int(vehicle_limits.get(name, 1000000))
                if limit > result[name]:
                    smallest_truck = name
                    break
            
            if smallest_truck:
                result[smallest_truck] += 1
                remaining = 0
            else:
                break
                
    return {k: v for k, v in result.items() if v > 0}


def get_vehicles_round_down(flow, vehicle_limits, caps=None, **kwargs):
    if caps is None:
        caps = {'7 L': 7000.0, '10 L': 10000.0, '12L': 12000.0, '15 L': 15000.0, '18 L': 18000.0, 'V30': 35000.0, 'V35': 35000.0}
    if isinstance(vehicle_limits, dict) and 'limits' in vehicle_limits:
        vehicle_limits = vehicle_limits['limits']
    sorted_caps = sorted(caps.items(), key=lambda x: x[1], reverse=True)

    def greedy_fallback():
        result = {k: 0 for k in caps}
        remaining = flow
        for cap_name, cap_val in sorted_caps:
            limit = int(vehicle_limits.get(cap_name, 1000000))
            if limit <= 0 or cap_val > remaining:
                continue
            taken = min(int(remaining // cap_val), limit)
            result[cap_name] = taken
            remaining -= taken * cap_val
        return {k: v for k, v in result.items() if v > 0}

    try:
        from ortools.linear_solver import pywraplp
        solver = pywraplp.Solver.CreateSolver('SCIP')
        if not solver:
            return greedy_fallback()

        vars_dict = {}
        for name, cap in caps.items():
            limit = int(vehicle_limits.get(name, 1000000))
            vars_dict[name] = solver.IntVar(0, limit, f"rd_{name}")

        total_cap_expr = sum(vars_dict[n] * c for n, c in caps.items())
        total_veh_expr = sum(vars_dict[n] for n in caps)
        
        solver.Add(total_cap_expr <= flow)
        
        # Always maximize capacity as priority 1, but minimize vehicles as priority 2 to break ties
        solver.Maximize(100.0 * total_cap_expr - 1.0 * total_veh_expr)
        
        status = solver.Solve()
        if status in (pywraplp.Solver.OPTIMAL, pywraplp.Solver.FEASIBLE):
            result = {}
            for name in caps:
                val = int(round(vars_dict[name].solution_value()))
                if val > 0:
                    result[name] = val
            return result
        return greedy_fallback()
    except Exception:
        return greedy_fallback()


def parse_plant_bmc_mapping(excel_file_path):
    mapping = set()
    if not excel_file_path or not os.path.exists(excel_file_path):
        return mapping
    try:
        df = pd.read_excel(excel_file_path, sheet_name='Plant_BMC_Mapping')
        for _, row in df.iterrows():
            p_code = str(row.get('PlantCode', '')).strip()
            bmc_code = str(row.get('BMCCode', '')).strip()
            commodity = str(row.get('commodity', '')).strip()
            if p_code and bmc_code and commodity:
                mapping.add((p_code, bmc_code, commodity))
    except Exception as e:
        print(f"Error parsing Plant_BMC_Mapping (may not exist): {e}")
    return mapping

def parse_bmc_vehicles(excel_file_path):
    vehicle_limits_map = {}
    if not excel_file_path or not os.path.exists(excel_file_path):
        vehicle_limits_map['global_caps'] = {'7 L': 7000.0, '10 L': 10000.0, '12L': 12000.0, '15 L': 15000.0, '18 L': 18000.0, 'V30': 35000.0, 'V35': 35000.0}
        return vehicle_limits_map
        
    try:
        xl = pd.ExcelFile(excel_file_path)
        
        # 1. Parse Vehicle Type to get dynamic capacities
        caps = {}
        caps_ranges = {}
        if 'Vehicle Type' in xl.sheet_names:
            df_vt = xl.parse('Vehicle Type')
            capacity_col = next((c for c in df_vt.columns if 'name' in str(c).lower() or 'capacity' in str(c).lower() or 'type' in str(c).lower()), None)
            code_col = next((c for c in df_vt.columns if 'code' in str(c).lower() or 'vehiclecode' in str(c).lower()), 'VehicleCode')
            to_col = next((c for c in df_vt.columns if str(c).lower() == 'to'), 'To')
            from_col = next((c for c in df_vt.columns if str(c).lower() == 'from'), 'From')
            
            if code_col in df_vt.columns and to_col in df_vt.columns:
                for _, row in df_vt.iterrows():
                    vc = str(row[code_col]).strip()
                    to_val = float(row[to_col]) if pd.notnull(row[to_col]) else 0.0
                    from_val = float(row[from_col]) if from_col in df_vt.columns and pd.notnull(row[from_col]) else (to_val * 0.8)
                    caps[vc] = to_val * 1000.0
                    caps_ranges[vc] = (from_val * 1000.0, to_val * 1000.0)
                    
                    # Fix for V35: The user set 'To' = 5000 in Excel as a hack to mean "unlimited upper bound"
                    # for the flow filter, but this makes the physical vehicle capacity 5,000,000 L.
                    # This breaks the 15% empty space rule (e.g. 37,180 flow leaves ~4.96M empty space).
                    # We restrict V35's physical capacity to a realistic 40,000 L (40 KL tanker).
                    if vc == 'V35' and caps[vc] > 100000:
                        caps[vc] = 40000.0
                        caps_ranges[vc] = (30000.0, 40000.0)
        
        # Fallback
        if not caps:
            caps = {'7 L': 7000.0, '10 L': 10000.0, '12L': 12000.0, '15 L': 15000.0, '18 L': 18000.0, 'V30': 35000.0, 'V35': 35000.0}
        if not caps_ranges:
            caps_ranges = {'7 L': (6000.0, 7000.0), '10 L': (8000.0, 10000.0), '12L': (10000.0, 12000.0), '15 L': (12000.0, 15000.0), '18 L': (15000.0, 18000.0), 'V30': (25000.0, 35000.0), 'V35': (30000.0, 35000.0)}
            
        if 'V30' not in caps:
            caps['V30'] = 35000.0
            caps_ranges['V30'] = (25000.0, 35000.0)
        if 'V35' not in caps:
            caps['V35'] = 35000.0
            caps_ranges['V35'] = (30000.0, 35000.0)
            
        vehicle_limits_map['global_caps'] = caps
        vehicle_limits_map['global_caps_ranges'] = caps_ranges

        sheet_name = None
        for s in xl.sheet_names:
            if s.lower().replace(' ', '').replace('_', '') == 'vehiclesupplierallocation':
                sheet_name = s
                break
        if not sheet_name:
            for s in xl.sheet_names:
                if 'supplier' in s.lower() and 'allocation' in s.lower():
                    sheet_name = s
                    break
        if sheet_name:
            df_veh = xl.parse(sheet_name)
            # Find ID column
            id_col = next((c for c in df_veh.columns if str(c).lower().replace(' ', '').replace('_', '') == 'suppliercluster'), None)
            if not id_col and len(df_veh.columns) > 0:
                id_col = df_veh.columns[0]
                
            if id_col:
                col_strat = next((c for c in df_veh.columns if 'strategy' in str(c).lower().replace(' ', '').replace('_', '')), None)
                col_margin_low = next((c for c in df_veh.columns if 'flowlowmargin' in str(c).lower().replace(' ', '').replace('_', '')), None)
                col_margin_high = next((c for c in df_veh.columns if 'flowhighmargin' in str(c).lower().replace(' ', '').replace('_', '')), None)
                col_cluster = next((c for c in df_veh.columns if 'cluster' in str(c).lower() and 'sub' not in str(c).lower()), None)
                col_subcluster = next((c for c in df_veh.columns if 'subcluster' in str(c).lower().replace(' ', '').replace('_', '')), None)
                
                vehicle_cols = {}
                for vc in caps.keys():
                    vc_norm = str(vc).lower().replace(' ', '').strip()
                    for c in df_veh.columns:
                        if str(c).lower().replace(' ', '').strip() == vc_norm:
                            vehicle_cols[vc] = c
                            break
                            
                for _, row in df_veh.iterrows():
                    bmc_id = str(row[id_col]).strip()
                    if bmc_id and bmc_id.lower() != 'nan':
                        limits = {}
                        for vc in caps.keys():
                            c_name = vehicle_cols.get(vc)
                            limits[vc] = float(row[c_name]) if (c_name and pd.notna(row[c_name])) else 1000000.0
                            
                        strategy = str(row[col_strat]).strip() if col_strat and pd.notna(row[col_strat]) else "Least Vehicle Strategy"
                        margin_low = float(row[col_margin_low]) if col_margin_low and pd.notna(row[col_margin_low]) else 5.0
                        margin_high = float(row[col_margin_high]) if col_margin_high and pd.notna(row[col_margin_high]) else 5.0
                        cluster = str(row[col_cluster]).strip() if col_cluster and pd.notna(row[col_cluster]) else ""
                        subcluster = str(row[col_subcluster]).strip() if col_subcluster and pd.notna(row[col_subcluster]) else ""
                        
                        vehicle_limits_map[bmc_id] = {
                            'limits': limits,
                            'strategy': strategy,
                            'margin_low': margin_low,
                            'margin_high': margin_high,
                            'cluster': cluster,
                            'subcluster': subcluster
                        }
            print(f"Loaded vehicle limits for {len(vehicle_limits_map)-1} Suppliers from Excel")
    except Exception as e:
        print("Error parsing BMCVechicle sheet:", e)
        
    return vehicle_limits_map


def solve_network_lp(hubs, plants, transport_cost_per_km=0.005, excel_file_path=None):
    plant_bmc_mapping = parse_plant_bmc_mapping(excel_file_path)
    try:
        from ortools.linear_solver import pywraplp
    except ImportError:
        return {
            'status': 'ERROR',
            'message': 'Google OR-Tools is not installed or available in the environment.'
        }

    # Load BMC vehicle limits from uploaded Excel file if it exists
    vehicle_limits_map = parse_bmc_vehicles(excel_file_path)
    
    # Create BMC to Supplier Mapping for quick lookup
    bmc_to_supplier = {}
    try:
        if excel_file_path and os.path.exists(excel_file_path):
            df_map = pd.read_excel(excel_file_path, sheet_name='Plant_BMC_Mapping')
            if 'Supplier' in df_map.columns:
                df_map['Supplier'] = df_map['Supplier'].astype(str).str.strip()
                bmc_to_supplier = df_map.drop_duplicates(subset=['BMCCode']).set_index('BMCCode')['Supplier'].to_dict()
    except:
        pass

    # Initialize solver
    solver = pywraplp.Solver.CreateSolver('GLOP')
    if not solver:
        return {'status': 'ERROR', 'message': 'Could not create GLOP solver.'}

    # Pre-calculate candidate distances using OSRM Table API
    total_nodes_count = len(hubs) + len(plants)
    is_large = total_nodes_count > 50
    dist_cache = {}
    
    # Load precalculated distances from Distance sheet in Excel if present
    if excel_file_path and os.path.exists(excel_file_path):
        try:
            xl = pd.ExcelFile(excel_file_path)
            distance_sheet_name = None
            for s in xl.sheet_names:
                if s.lower().strip() == 'distance':
                    distance_sheet_name = s
                    break
            if distance_sheet_name:
                df_dist = xl.parse(distance_sheet_name)
                col_hub = next((c for c in df_dist.columns if any(k in str(c).lower() for k in ['hub', 'bmc', 'mcc', 'source', 'from'])), None)
                col_plant = next((c for c in df_dist.columns if any(k in str(c).lower() for k in ['plant', 'to', 'dest', 'target'])), None)
                col_dist = next((c for c in df_dist.columns if any(k in str(c).lower() for k in ['dist', 'km'])), None)
                
                if col_hub and col_plant and col_dist:
                    def canonicalize_id(val):
                        if pd.isna(val) or val is None:
                            return ""
                        return "".join(c for c in str(val).lower() if c.isalnum())
                    
                    hub_id_map = {}
                    for h in hubs:
                        hub_id_map[canonicalize_id(h['id'])] = h['id']
                        hub_id_map[canonicalize_id(h.get('name'))] = h['id']
                        
                    plant_id_map = {}
                    for p in plants:
                        plant_id_map[canonicalize_id(p['id'])] = p['id']
                        plant_id_map[canonicalize_id(p.get('name'))] = p['id']
                        
                    loaded_count = 0
                    for _, row in df_dist.iterrows():
                        h_raw = canonicalize_id(row[col_hub])
                        p_raw = canonicalize_id(row[col_plant])
                        d_val = row[col_dist]
                        
                        h_id = hub_id_map.get(h_raw)
                        p_id = plant_id_map.get(p_raw)
                        
                        if h_id and p_id and pd.notna(d_val):
                            dist_cache[(h_id, p_id)] = float(d_val)
                            pair_key = tuple(sorted([h_id, p_id]))
                            _distance_cache[pair_key] = float(d_val)
                            loaded_count += 1
                            
                    print(f"Loaded {loaded_count} precalculated distances from Excel sheet 'Distance'")
        except Exception as e:
            print("Error loading distances from Excel sheet 'Distance':", e)

    # 1. Hub -> Plant
    for h in hubs:
        for p in plants:
            if (h['id'], p['id']) not in dist_cache:
                dist_cache[(h['id'], p['id'])] = 9999.0
                


    # Helper to resolve distance
    def get_pair_dist(node1, node2):
        key = (node1['id'], node2['id'])
        if key in dist_cache:
            return dist_cache[key]
        return 9999.0

    # Dynamically extract all raw milk types present in hubs
    milk_types = set()
    for h in hubs:
        h_prods = h.get('products', [])
        if not h_prods and 'capacity' in h:
            h_prods = [{'type': 'Cow Milk', 'capacity': h['capacity'], 'processing_cost': h.get('processing_cost', 0)}]
        for p in h_prods:
            if 'type' in p:
                milk_types.add(p['type'])
    if not milk_types:
        milk_types = {'Cow Milk', 'Buffalo Milk'}
    milk_types = sorted(list(milk_types))

    # Decision Variables
    # Flow: Hub -> Plant per Milk Type
    flow_h_p = {}
    for h in hubs:
        h_prods = h.get('products', [])
        if not h_prods and 'capacity' in h:
            h_prods = [{'type': 'Cow Milk', 'capacity': h['capacity']}]
        h_types = {p['type'] for p in h_prods if 'type' in p}
        for p in plants:
            if is_large and (h['id'], p['id']) not in dist_cache:
                continue
            
            p_demands = p.get('demands', [])
            p_demand_types = {d['type'] for d in p_demands if 'type' in d}
            
            p_products = p.get('products', [])
            if not p_products and 'production_type' in p:
                p_products = [{'type': p['production_type']}]
            p_milk_types = {get_milk_type_for_product(prod['type']) for prod in p_products if 'type' in prod}
            
            common_milk = h_types.intersection(p_milk_types.union(p_demand_types))
            if common_milk:
                dist = get_pair_dist(h, p)
                if dist <= MAX_DISTANCE_LIMIT:
                    for m in sorted(common_milk):
                        # Strict mapping logic: if mapping exists, only allow mapped routes
                        if plant_bmc_mapping and (str(p['id']), str(h['id']), str(m)) not in plant_bmc_mapping:
                            continue
                            
                        clean_m = m.replace(' ', '_').replace('-', '_')
                        name = f"flow_H_{h['id']}_P_{p['id']}_{clean_m}"
                        flow_h_p[(h['id'], p['id'], m)] = solver.NumVar(0, solver.infinity(), name)

    # Precompute nearest plant map for each hub and milk type
    nearest_plant_map = {}
    for h in hubs:
        dev_vars = []
        for m in milk_types:
            compatible_plants = [p for p in plants if (h['id'], p['id'], m) in flow_h_p]
            if compatible_plants:
                p_near = min(compatible_plants, key=lambda p: get_pair_dist(h, p))
                nearest_plant_map[(h['id'], m)] = p_near['id']

    # Required vs Excess Flow Variables for Hub -> Plant
    flow_required = {}
    flow_excess = {}
    for (h_id, p_id, m), flow_var in flow_h_p.items():
        clean_m = m.replace(' ', '_').replace('-', '_')
        flow_required[(h_id, p_id, m)] = solver.NumVar(0, solver.infinity(), f"flow_req_H_{h_id}_P_{p_id}_{clean_m}")
        flow_excess[(h_id, p_id, m)] = solver.NumVar(0, solver.infinity(), f"flow_exc_H_{h_id}_P_{p_id}_{clean_m}")
        solver.Add(flow_var == flow_required[(h_id, p_id, m)] + flow_excess[(h_id, p_id, m)])
        
        # If this plant is not the nearest plant for this hub/milk, excess flow must be 0
        p_near_id = nearest_plant_map.get((h_id, m))
        if p_near_id and p_id != p_near_id:
            solver.Add(flow_excess[(h_id, p_id, m)] == 0)

    # Constraints
    slack_vars = []

    # Precompute plant inflows for the Even Distribution Rule (using required flow only)
    plant_inflow_vars = {}
    for p in plants:
        dev_vars = []
        for m in milk_types:
            plant_inflow_vars[(p['id'], m)] = [
                flow_required[(h_other['id'], p['id'], m)]
                for h_other in hubs
                if (h_other['id'], p['id'], m) in flow_h_p
            ]

    # 1. Hub supply full milk limits (per milk type)
    for h in hubs:
        h_prods = h.get('products', [])
        if not h_prods and 'capacity' in h:
            h_prods = [{'type': 'Cow Milk', 'capacity': h['capacity'], 'processing_cost': h.get('processing_cost', 0)}]
        capacity_dict = {p['type']: p.get('capacity', 0) for p in h_prods if 'type' in p}
        
        dev_vars = []
        for m in milk_types:
            flow_out_vars = [flow_h_p[(h['id'], p['id'], m)] for p in plants if (h['id'], p['id'], m) in flow_h_p]
            cap_limit = capacity_dict.get(m, 0)
            if cap_limit > 0:
                if flow_out_vars:
                    slack = solver.NumVar(0, solver.infinity(), f"slack_hub_{h['id']}_{m.replace(' ', '_')}")
                    solver.Add(sum(flow_out_vars) + slack == cap_limit)
                    slack_vars.append(slack * 100000.0)
                    
                    # [REMOVED] Even distribution is now handled in Stage 2 lexicographic optimization.

            else:
                if flow_out_vars:
                    solver.Add(sum(flow_out_vars) == 0)

    # 2. Plant capacity & flow conservation (conservation per milk type, total capacity limit)
    for p in plants:
        p_products = p.get('products', [])
        if not p_products and 'production_type' in p:
            p_products = [{'type': p['production_type']}]
        
        p_demands = p.get('demands', [])
        demand_dict = {d['type']: d.get('demand', 0) for d in p_demands if 'type' in d}
        
        inflow_milks = p.get('inflow_milks', [])
        if not inflow_milks:
            if p_demands:
                inflow_milks = [{'type': d['type'], 'capacity': d['demand'], 'processing_cost': d.get('processing_cost', 0.40)} for d in p_demands]
            else:
                inflow_milks = [
                    {'type': 'Cow Milk', 'capacity': p.get('capacity', 10000), 'processing_cost': p.get('processing_cost', 0.50)},
                    {'type': 'Buffalo Milk', 'capacity': p.get('capacity', 10000), 'processing_cost': p.get('processing_cost', 0.50)}
                ]
        capacity_dict = {m['type']: m.get('capacity', 0) for m in inflow_milks if 'type' in m}
        
        # Substitution variables: BM -> FCM -> MM
        trans_BM_FCM = solver.NumVar(0, solver.infinity(), f"trans_{p['id']}_BM_to_FCM")
        trans_FCM_MM = solver.NumVar(0, solver.infinity(), f"trans_{p['id']}_FCM_to_MM")
        
        if 'trans_vars' not in locals():
            trans_vars = {}
        trans_vars[(p['id'], 'BM_to_FCM')] = trans_BM_FCM
        trans_vars[(p['id'], 'FCM_to_MM')] = trans_FCM_MM
        
        # Small penalty to prefer direct matches before downgrading
        slack_vars.append(trans_BM_FCM * 10.0)
        slack_vars.append(trans_FCM_MM * 10.0)
        
        dev_vars = []
        for m in milk_types:
            flow_in_m = sum(flow_h_p[(h['id'], p['id'], m)] for h in hubs if (h['id'], p['id'], m) in flow_h_p)
            
            effective_flow_in = flow_in_m
            m_lower = str(m).lower()
            if 'buffalo' in m_lower or 'bm' in m_lower:
                effective_flow_in = flow_in_m - trans_BM_FCM
            elif 'fcm' in m_lower:
                effective_flow_in = flow_in_m + trans_BM_FCM - trans_FCM_MM
            elif 'mm' in m_lower:
                effective_flow_in = flow_in_m + trans_FCM_MM
            
            # Bound required flow into plant to its required supply limit (R_{p,m})
            r_limit = 0.0
            if m in demand_dict:
                r_limit = demand_dict[m]
            else:
                r_limit = capacity_dict.get(m, 0.0)
            
            req_inflow_vars = [
                flow_required[(h['id'], p['id'], m)]
                for h in hubs
                if (h['id'], p['id'], m) in flow_h_p
            ]
            if req_inflow_vars:
                solver.Add(sum(req_inflow_vars) <= r_limit)
            
            cap_limit = demand_dict.get(m, capacity_dict.get(m, 0))
            if cap_limit > 0:
                # Soft capacity limit: effective_flow_in <= cap_limit + over_cap
                over_cap = solver.NumVar(0, solver.infinity(), f"over_cap_{p['id']}_{m.replace(' ', '_')}")
                solver.Add(effective_flow_in - over_cap <= cap_limit)
                slack_vars.append(over_cap * 1000.0)
            else:
                solver.Add(effective_flow_in == 0)

        # Ensure plant is fulfilled at least 20 percent of its capacity
        total_capacity = sum(capacity_dict.values()) if capacity_dict else p.get('capacity', 0)
        if total_capacity > 0:
            inflow_vars = [flow_h_p[(h['id'], p['id'], m)] for h in hubs for m in milk_types if (h['id'], p['id'], m) in flow_h_p]
            if inflow_vars:
                plant_slack = solver.NumVar(0, solver.infinity(), f"slack_plant_{p['id']}")
                solver.Add(sum(inflow_vars) + plant_slack >= 0.20 * total_capacity)
                slack_vars.append(plant_slack * 1000.0)


    # Objective: Maximize Profit
    # Profit = Revenue - Transport Cost - Processing/Handling Cost

    # 1. Revenue
    revenue_items = []

    for p in plants:
        p_demands = p.get('demands', [])
        for d in p_demands:
            m = d['type']
            price = d.get('price', 0.0)
            flow_in_m_vars = [flow_h_p[(h['id'], p['id'], m)] for h in hubs if (h['id'], p['id'], m) in flow_h_p]
            if flow_in_m_vars:
                revenue_items.append(sum(flow_in_m_vars) * price)
                
    revenue_expr = sum(revenue_items)

    # 2. Transport Costs
    trans_cost_expr = []
    # Hub -> Plant
    for (h_id, p_id, m), flow_var in flow_h_p.items():
        h_node = next(x for x in hubs if x['id'] == h_id)
        p_node = next(x for x in plants if x['id'] == p_id)
        dist = get_pair_dist(h_node, p_node)
        # Use actual distance to favor nearest plant
        effective_dist = dist
        cost_per_unit = effective_dist * transport_cost_per_km
        
        # Commodity Priority Logic: BM > FCM > MM
        m_lower = str(m).lower()
        bonus = 0.0
        if 'buffalo' in m_lower or 'bm' in m_lower:
            bonus = transport_cost_per_km * 0.6  # Equivalent to 0.003 when cost is 0.005
        elif 'fcm' in m_lower:
            bonus = transport_cost_per_km * 0.4  # Equivalent to 0.002 when cost is 0.005
        elif 'mm' in m_lower:
            bonus = transport_cost_per_km * 0.2  # Equivalent to 0.001 when cost is 0.005
            
        cost_per_unit -= bonus
            
        trans_cost_expr.append(flow_var * cost_per_unit)


    # 3. Processing/Handling Costs at Hubs and Plants
    proc_cost_expr = []
    for h in hubs:
        h_prods = h.get('products', [])
        if not h_prods and 'capacity' in h:
            h_prods = [{'type': 'Cow Milk', 'capacity': h['capacity'], 'processing_cost': h.get('processing_cost', 0)}]
        cost_dict = {p['type']: p.get('processing_cost', 0.0) for p in h_prods if 'type' in p}
        dev_vars = []
        for m in milk_types:
            flow_out_m = sum(flow_h_p[(h['id'], p['id'], m)] for p in plants if (h['id'], p['id'], m) in flow_h_p)
            proc_cost_expr.append(flow_out_m * cost_dict.get(m, 0.0))
            
    for p in plants:
        inflow_milks = p.get('inflow_milks', [])
        if not inflow_milks:
            p_demands = p.get('demands', [])
            if p_demands:
                inflow_milks = [{'type': d['type'], 'capacity': d['demand'], 'processing_cost': d.get('processing_cost', 0.40)} for d in p_demands]
            else:
                inflow_milks = [
                    {'type': 'Cow Milk', 'capacity': p.get('capacity', 10000), 'processing_cost': p.get('processing_cost', 0.50)},
                    {'type': 'Buffalo Milk', 'capacity': p.get('capacity', 10000), 'processing_cost': p.get('processing_cost', 0.50)}
                ]
        cost_dict = {m['type']: m.get('processing_cost', 0.0) for m in inflow_milks if 'type' in m}
        dev_vars = []
        for m in milk_types:
            flow_in_m = sum(flow_h_p[(h['id'], p['id'], m)] for h in hubs if (h['id'], p['id'], m) in flow_h_p)
            proc_cost_expr.append(flow_in_m * cost_dict.get(m, 0.40))
    # Stage 1: Distance and Profit Optimization
    slack_total = solver.NumVar(0, solver.infinity(), 'slack_total')
    solver.Add(slack_total == sum(slack_vars))
    
    trans_total = solver.NumVar(0, solver.infinity(), 'trans_total')
    solver.Add(trans_total == sum(trans_cost_expr))
    
    rev_total = solver.NumVar(-solver.infinity(), solver.infinity(), 'rev_total')
    solver.Add(rev_total == revenue_expr)
    
    profit_expr = rev_total - trans_total - sum(proc_cost_expr) - slack_total
    solver.Maximize(profit_expr)
    status = solver.Solve()

    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        opt_slack = slack_total.solution_value()
        opt_rev = rev_total.solution_value()
        opt_trans = trans_total.solution_value()
        
        # Stage 2: Fair Rebalancing (Even Distribution)
        # 1. Lock slack penalties so the solver doesn't drop milk to cheat
        solver.Add(slack_total <= opt_slack + 0.1)
        # 2. Lock revenue so we don't lose sales
        solver.Add(rev_total >= opt_rev - 0.1)
        # 3. Add Tolerance Budget for Transport: Allow up to MAX_DISTANCE_LIMIT implications
        # By giving transport a massive generous budget (or not locking it), the solver is free 
        # to route milk to ANY valid mapped plant within MAX_DISTANCE_LIMIT to achieve Even Distribution.
        solver.Add(trans_total <= opt_trans + (opt_trans * 5.0) + 50000.0) # 500% tolerance budget!
        
        dev_vars = []
        for m in milk_types:
            target_F_m = solver.NumVar(0, solver.infinity(), f'target_F_{m.replace(" ", "_")}')
            max_dev_m = solver.NumVar(0, solver.infinity(), f'max_dev_{m.replace(" ", "_")}')
            has_plants_for_m = False
            
            m_lower = str(m).lower()
            
            for p in plants:
                inflows = [flow_h_p[(h['id'], p['id'], m)] for h in hubs if (h['id'], p['id'], m) in flow_h_p]
                if not inflows: continue
                total_inflow = sum(inflows)
                
                # Determine capacity for this specific commodity
                cap = 0
                p_demands = p.get('demands', [])
                if p_demands:
                    demand_dict = {d['type']: d.get('demand', 0) for d in p_demands if 'type' in d}
                    if m in demand_dict:
                        cap = demand_dict[m]
                    elif 'cow' in m_lower or 'cm' in m_lower:
                        demand_keys = [k for k in demand_dict.keys() if 'cow' in k.lower() or 'cm' in k.lower()]
                        if demand_keys: cap = sum(demand_dict[k] for k in demand_keys)
                    elif 'buffalo' in m_lower or 'bm' in m_lower:
                        demand_keys = [k for k in demand_dict.keys() if 'buffalo' in k.lower() or 'bm' in k.lower()]
                        if demand_keys: cap = sum(demand_dict[k] for k in demand_keys)
                else:
                    inflow_milks = p.get('inflow_milks', [])
                    if inflow_milks:
                        capacity_dict = {im['type']: im.get('capacity', 0) for im in inflow_milks if 'type' in im}
                        if m in capacity_dict:
                            cap = capacity_dict[m]
                    else:
                        cap = p.get('capacity', 0)
                
                if cap <= 0: continue
                has_plants_for_m = True
                
                F_p = total_inflow * (1000.0 / cap)
                
                dev = solver.NumVar(0, solver.infinity(), f"dev_P_{p['id']}_{m.replace(' ', '_')}")
                solver.Add(dev >= F_p - target_F_m)
                solver.Add(dev >= target_F_m - F_p)
                
                # Min-max formulation: max_dev must be >= every plant's deviation for this commodity
                solver.Add(max_dev_m >= dev)
                
            if has_plants_for_m:
                dev_vars.append(max_dev_m)
                
        if dev_vars:
            # Minimize deviation as primary goal, but also minimize transport cost as a secondary goal
            # This ensures that even when perfectly balanced, the solver strictly prefers the nearest plants
            solver.Maximize(-sum(dev_vars) * 1000000.0 - trans_total)
            status = solver.Solve()

    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        routes = []

        # Pre-calculate actual substitutions per plant
        plant_subs = {}
        for p in plants:
            pid = p['id']
            bm_to_fcm = trans_vars[(pid, 'BM_to_FCM')].solution_value() if 'trans_vars' in locals() and (pid, 'BM_to_FCM') in trans_vars else 0.0
            fcm_to_mm = trans_vars[(pid, 'FCM_to_MM')].solution_value() if 'trans_vars' in locals() and (pid, 'FCM_to_MM') in trans_vars else 0.0
            
            # calculate inflows
            bm_inflow = sum(flow_h_p[(h['id'], pid, _m)].solution_value() for h in hubs for _m in milk_types if (h['id'], pid, _m) in flow_h_p and ('buffalo' in str(_m).lower() or 'bm' in str(_m).lower()))
            fcm_inflow = sum(flow_h_p[(h['id'], pid, _m)].solution_value() for h in hubs for _m in milk_types if (h['id'], pid, _m) in flow_h_p and 'fcm' in str(_m).lower())
            
            fcm_cascaded = min(fcm_to_mm, fcm_inflow) if fcm_inflow > 0 else 0.0
            bm_cascaded_to_mm = max(0, fcm_to_mm - fcm_inflow)
            bm_cascaded_to_fcm = max(0, bm_to_fcm - bm_cascaded_to_mm)
            
            plant_subs[pid] = {
                'BM_inflow': bm_inflow,
                'FCM_inflow': fcm_inflow,
                'BM_to_FCM_actual': bm_cascaded_to_fcm,
                'BM_to_MM_actual': bm_cascaded_to_mm,
                'FCM_to_MM_actual': fcm_cascaded
            }

        # Extract Hub -> Plant flows
        for (h_id, p_id, m), flow_var in flow_h_p.items():
            val = round(flow_var.solution_value())
            if val > 0:
                h_node = next(x for x in hubs if x['id'] == h_id)
                p_node = next(x for x in plants if x['id'] == p_id)
                dist = get_pair_dist(h_node, p_node)
                cost = dist * transport_cost_per_km
                clean_m = m.replace(' ', '_').replace('-', '_')
                
                # Calculate optimal vehicles
                supplier = bmc_to_supplier.get(h_id, '')
                limits = vehicle_limits_map.get(supplier, {})
                optimal_veh = get_optimal_vehicles(val, limits, caps=vehicle_limits_map.get('global_caps'), caps_ranges=vehicle_limits_map.get('global_caps_ranges'))
                global_caps = vehicle_limits_map.get('global_caps', {})
                total_veh = sum(optimal_veh.values())
                total_cap = sum(count * global_caps.get(v, 0) for v, count in optimal_veh.items())
                # Cap the flow at total_cap if we left some undispatched, but only if we dispatched at least one vehicle
                # (Removed overwrite of val to keep original solver flow)
                dispatch_qty = val
                if total_veh == 0:
                    dispatch_qty = 0.0
                elif 0 < total_cap < val:
                    dispatch_qty = total_cap
                excess = total_cap - dispatch_qty if total_veh > 0 else 0.0
                
                # Calculate trip-based cost
                vehicle_rates = {'V07': 38, 'V10': 42, 'V12': 46, 'V15': 52, 'V20': 60, 'V25': 68, 'V30': 75, 'V35': 85}
                cost = sum(trips * dist * vehicle_rates.get(v_type, 0) for v_type, trips in optimal_veh.items())
                
                primary_flow = float(val)
                extra_flows = []
                subs = plant_subs.get(p_id, {})
                m_lower = str(m).lower()
                
                if ('buffalo' in m_lower or 'bm' in m_lower) and subs.get('BM_inflow', 0) > 0:
                    prop = val / subs['BM_inflow']
                    r_bm_to_fcm = prop * subs['BM_to_FCM_actual']
                    r_bm_to_mm = prop * subs['BM_to_MM_actual']
                    
                    if plant_bmc_mapping and (str(p_id), str(h_id), "FCM") not in plant_bmc_mapping:
                        r_bm_to_fcm = 0.0
                    if plant_bmc_mapping and (str(p_id), str(h_id), "MM") not in plant_bmc_mapping:
                        r_bm_to_mm = 0.0
                        
                    primary_flow -= (r_bm_to_fcm + r_bm_to_mm)
                    if round(r_bm_to_fcm, 2) > 0:
                        extra_flows.append((r_bm_to_fcm, "BM to FCM", "FCM"))
                    if round(r_bm_to_mm, 2) > 0:
                        extra_flows.append((r_bm_to_mm, "BM to MM", "MM"))
                        
                elif 'fcm' in m_lower and subs.get('FCM_inflow', 0) > 0:
                    prop = val / subs['FCM_inflow']
                    r_fcm_to_mm = prop * subs['FCM_to_MM_actual']
                    
                    if plant_bmc_mapping and (str(p_id), str(h_id), "MM") not in plant_bmc_mapping:
                        r_fcm_to_mm = 0.0
                        
                    primary_flow -= r_fcm_to_mm
                    if round(r_fcm_to_mm, 2) > 0:
                        extra_flows.append((r_fcm_to_mm, "FCM to MM", "MM"))

                if round(primary_flow, 2) > 0 or not extra_flows:
                    route_dict = {
                        'id': f"route_{h_id}_{p_id}_{clean_m}",
                        'from_id': h_id,
                        'to_id': p_id,
                        'from_type': 'hub',
                        'to_type': 'plant',
                        'flow': round(primary_flow, 2),
                        'product_type': m,
                        'target_product_type': m,
                        'unit': 'L',
                        'distance': round(dist, 2),
                        'cost': round(cost, 2),
                        'total_vehicles': total_veh,
                        'total_vehicle_capacity': total_cap,
                        'excess_vehicle_capacity': round(excess, 2)
                    }
                    for v_type, trips in optimal_veh.items():
                        route_dict[f'vehicles_{v_type}'] = trips
                    routes.append(route_dict)
                    
                for ext_flow, ext_ptype, target_group in extra_flows:
                    ext_route_dict = {
                        'id': f"route_{h_id}_{p_id}_{clean_m}_extra_{ext_ptype.replace(' ', '_')}",
                        'from_id': h_id,
                        'to_id': p_id,
                        'from_type': 'hub',
                        'to_type': 'plant',
                        'flow': round(ext_flow, 2),
                        'product_type': ext_ptype,
                        'target_product_type': target_group,
                        'unit': 'L',
                        'distance': round(dist, 2),
                        'cost': 0.0,
                        'total_vehicles': 0,
                        'total_vehicle_capacity': 0,
                        'excess_vehicle_capacity': 0.0
                    }
                    for v_type in vehicle_rates.keys():
                        ext_route_dict[f'vehicles_{v_type}'] = 0
                    routes.append(ext_route_dict)


        # Aggregate metrics
        obj_val = solver.Objective().Value()
        
        total_revenue = 0.0

        for p in plants:
            p_demands = p.get('demands', [])
            for d in p_demands:
                m = d['type']
                price = d.get('price', 0.0)
                flow_in_val = sum(flow_h_p[(h['id'], p['id'], m)].solution_value() for h in hubs if (h['id'], p['id'], m) in flow_h_p)
                if flow_in_val > 0.1:
                    total_revenue += flow_in_val * price

        total_transport_cost = sum(r['cost'] for r in routes)

        total_hub_proc_cost = 0.0
        for h in hubs:
            h_prods = h.get('products', [])
            if not h_prods and 'capacity' in h:
                h_prods = [{'type': 'Cow Milk', 'capacity': h['capacity'], 'processing_cost': h.get('processing_cost', 0)}]
            cost_dict = {p['type']: p.get('processing_cost', 0.0) for p in h_prods if 'type' in p}
            dev_vars = []
        for m in milk_types:
                flow_out_m = sum(flow_h_p[(h['id'], p['id'], m)].solution_value() for p in plants if (h['id'], p['id'], m) in flow_h_p)
                total_hub_proc_cost += flow_out_m * cost_dict.get(m, 0.0)

        total_plant_proc_cost = 0.0
        for p in plants:
            inflow_milks = p.get('inflow_milks', [])
            if not inflow_milks:
                p_demands = p.get('demands', [])
                if p_demands:
                    inflow_milks = [{'type': d['type'], 'capacity': d['demand'], 'processing_cost': d.get('processing_cost', 0.40)} for d in p_demands]
                else:
                    inflow_milks = [
                        {'type': 'Cow Milk', 'capacity': p.get('capacity', 10000), 'processing_cost': p.get('processing_cost', 0.50)},
                        {'type': 'Buffalo Milk', 'capacity': p.get('capacity', 10000), 'processing_cost': p.get('processing_cost', 0.50)}
                    ]
            cost_dict = {m['type']: m.get('processing_cost', 0.0) for m in inflow_milks if 'type' in m}
            dev_vars = []
        for m in milk_types:
                flow_in_m = sum(flow_h_p[(h['id'], p['id'], m)].solution_value() for h in hubs if (h['id'], p['id'], m) in flow_h_p)
                total_plant_proc_cost += flow_in_m * cost_dict.get(m, 0.40)

        total_proc_cost = total_hub_proc_cost + total_plant_proc_cost

        # Flow summary per node
        node_metrics = {}
        for h in hubs:
            out_val = sum(flow_h_p[(h['id'], p['id'], m)].solution_value() for p in plants for m in milk_types if (h['id'], p['id'], m) in flow_h_p)
            node_metrics[h['id']] = {
                'inflow': 0.0,
                'outflow': round(out_val, 2)
            }
        for p in plants:
            in_val = sum(flow_h_p[(h['id'], p['id'], m)].solution_value() for h in hubs for m in milk_types if (h['id'], p['id'], m) in flow_h_p)
            
            outflow_milk = in_val

            node_metrics[p['id']] = {
                'inflow': round(in_val, 2),
                'outflow': round(outflow_milk, 2)
            }
        return {
            'status': 'OPTIMAL',
            'summary': {
                'profit': round(obj_val, 2),
                'revenue': round(total_revenue, 2),
                'transport_cost': round(total_transport_cost, 2),
                'processing_cost': round(total_proc_cost, 2),
                'hub_processing_cost': round(total_hub_proc_cost, 2),
                'plant_processing_cost': round(total_plant_proc_cost, 2)
            },
            'routes': routes,
            'node_metrics': node_metrics,
            'trans_vars_solution': {k: v.solution_value() for k, v in trans_vars.items()} if 'trans_vars' in locals() else {}
        }
    else:
        return {
            'status': 'INFEASIBLE',
            'message': 'The model is infeasible or unbounded. Check capacities, supplies, and product compatibility.'
        }


# Background thread processor for jobs
def process_job_in_background(job_id, network_id, nodes, transport_cost_per_km, excel_file_path=None):
    global MAX_DISTANCE_LIMIT
    MAX_DISTANCE_LIMIT = 800.0  # Reset to default for each job to prevent mixing data between jobs
    try:
        from src.config.environment import Environment
        import pymongo
        client = pymongo.MongoClient(Environment.MONGO_URI)
        setting = client[Environment.MONGO_DB]["RequestSettings"].find_one({"requestId": job_id})
        if setting and "maxDistance" in setting:
            MAX_DISTANCE_LIMIT = float(setting["maxDistance"])
        client.close()
    except Exception as e:
        print(f"Error fetching MAX_DISTANCE for job {job_id}: {e}")

    start_time = time.time()
    
    try:
        hubs = [n for n in nodes if n['type'] == 'hub']
        plants = [n for n in nodes if n['type'] == 'plant']
        
        # Run solver helper
        res = solve_network_lp(hubs, plants, transport_cost_per_km, excel_file_path)
        
        if res.get('status') in ('OPTIMAL', 'FEASIBLE'):
            vehicle_limits_map = parse_bmc_vehicles(excel_file_path)
            plant_bmc_mapping = parse_plant_bmc_mapping(excel_file_path)
            valid_route_tuples = {(str(bmc), str(plant), str(commodity)) for plant, bmc, commodity in plant_bmc_mapping}
            
            # Map BMCs to Suppliers (need to extract from mapping file if not available otherwise)
            bmc_to_supplier = {}
            if excel_file_path and os.path.exists(excel_file_path):
                try:
                    df_map = pd.read_excel(excel_file_path, sheet_name='Plant_BMC_Mapping')
                    if 'Supplier' in df_map.columns:
                        df_map['Supplier'] = df_map['Supplier'].astype(str).str.strip()
                        bmc_to_supplier = df_map.drop_duplicates(subset=['BMCCode']).set_index('BMCCode')['Supplier'].to_dict()
                except:
                    pass

            output_filename = f"results_{job_id}.xlsx"
            output_path = os.path.join(OUTPUT_FOLDER, output_filename)
            
            # Initialize sub-cluster vehicle pools
            subcluster_vehicle_pools = {}
            for supplier_id, info in vehicle_limits_map.items():
                if not isinstance(info, dict) or 'limits' not in info:
                    continue
                c = info.get('cluster', '')
                sc = info.get('subcluster', '')
                key = (c, sc)
                if key not in subcluster_vehicle_pools:
                    subcluster_vehicle_pools[key] = {vc: 0.0 for vc in vehicle_limits_map.get('global_caps', {}).keys()}
                limits = info['limits']
                for vc in vehicle_limits_map.get('global_caps', {}).keys():
                    subcluster_vehicle_pools[key][vc] += limits.get(vc, 0.0)

            # Find BMC capacities
            bmc_capacities = {}
            for h in hubs:
                h_prods = h.get('products', [])
                if not h_prods and 'capacity' in h:
                    h_prods = [{'type': 'Cow Milk', 'capacity': h['capacity']}]
                for p in h_prods:
                    if 'type' in p:
                        bmc_capacities[(h['id'], p['type'])] = p.get('capacity', 0.0)

            # Calculate total flow from each BMC for each milk type
            total_flow_from_bmc = {}
            for r in res.get('routes', []):
                if r.get('from_type') == 'hub':
                    key = (r['from_id'], r['product_type'])
                    total_flow_from_bmc[key] = total_flow_from_bmc.get(key, 0.0) + r['flow']

            # Separate Hub -> Plant routes
            hub_routes = [r for r in res.get('routes', []) if r.get('from_type') == 'hub']
            
            # Helper to get cluster and subcluster for sorting and grouping
            def get_subcluster_key(r):
                supplier = bmc_to_supplier.get(r['from_id'], '')
                bmc_info = vehicle_limits_map.get(supplier, {})
                c = bmc_info.get('cluster', '') if isinstance(bmc_info, dict) else ''
                sc = bmc_info.get('subcluster', '') if isinstance(bmc_info, dict) else ''
                return (c, sc)
                
            from collections import defaultdict
            grouped_hub_routes = defaultdict(list)
            for r in hub_routes:
                group_type = r.get('target_product_type', r['product_type'])
                grouped_hub_routes[(r['from_id'], r['to_id'], group_type)].append(r)
                
            group_keys = list(grouped_hub_routes.keys())
            group_keys.sort(key=lambda k: (get_subcluster_key(grouped_hub_routes[k][0]), -sum(r['flow'] for r in grouped_hub_routes[k])))

            # Process Hub -> Plant route groups using sub-cluster pools and supply rules
            for g_key in group_keys:
                group = grouped_hub_routes[g_key]
                group.sort(key=lambda r: 1 if '_extra_' in str(r.get('id', '')) else 0)
                r_main = group[0]
                
                c, sc = get_subcluster_key(r_main)
                key = (c, sc)
                
                # Retrieve remaining sub-cluster vehicle pool
                limits = subcluster_vehicle_pools.get(key, {
                    vc: 1000000.0 for vc in vehicle_limits_map.get('global_caps', {}).keys()
                })
                
                # Margins
                supplier = bmc_to_supplier.get(r_main['from_id'], '')
                bmc_info = vehicle_limits_map.get(supplier, {})
                margin_low = bmc_info.get('margin_low', 5.0) if isinstance(bmc_info, dict) else 5.0
                margin_high = bmc_info.get('margin_high', 5.0) if isinstance(bmc_info, dict) else 5.0
                strategy = bmc_info.get('strategy', 'Least Vehicle Strategy') if isinstance(bmc_info, dict) else 'Least Vehicle Strategy'
                
                q = sum(r['flow'] for r in group)
                
                # Left quantity on BMC
                left_qty = bmc_capacities.get((r_main['from_id'], r_main['product_type']), 0.0) - total_flow_from_bmc.get((r_main['from_id'], r_main['product_type']), 0.0)
                
                # Temporary vehicle allocation
                optimal_veh = get_optimal_vehicles(q, limits, caps=vehicle_limits_map.get('global_caps'), distance=r_main['distance'], strategy=strategy, margin=margin_high)
                global_caps = vehicle_limits_map.get('global_caps', {})
                total_veh = sum(optimal_veh.values())
                total_cap = sum(count * global_caps.get(v, 0) for v, count in optimal_veh.items())
                excess_qty = total_cap - q if total_veh > 0 else 0.0
                
                do_not_supply  = False
                reason_override = None
                per_veh_empty  = max(0.0, excess_qty)

                lq_val = None
                if optimal_veh:
                    global_caps = vehicle_limits_map.get('global_caps', {'7 L': 7000.0, '10 L': 10000.0, '12L': 12000.0, '15 L': 15000.0, '18 L': 18000.0, 'V30': 35000.0, 'V35': 35000.0})
                    smallest_cap = min((global_caps.get(k, 0) for k, v in optimal_veh.items() if v > 0), default=None)
                    if smallest_cap:
                        lq_val = 0.15 * smallest_cap

                if total_veh == 0 and q > 0.1:
                    do_not_supply = True
                    reason_override = "No vehicles available in the sub-cluster pool"

                elif lq_val is not None and excess_qty > lq_val and str(strategy).strip().lower() != 'least vehicle strategy':
                    rd_veh = get_vehicles_round_down(q, limits, caps=vehicle_limits_map.get('global_caps'), strategy=strategy)
                    global_caps = vehicle_limits_map.get('global_caps', {})
                    rd_total_veh = sum(rd_veh.values())
                    rd_total_cap = sum(count * global_caps.get(v, 0) for v, count in rd_veh.items())

                    if rd_total_veh > 0:
                        optimal_veh = rd_veh
                        total_veh   = rd_total_veh
                        total_cap   = rd_total_cap
                        excess_qty  = rd_total_cap - q
                        per_veh_empty = 0.0
                        deficit     = q - rd_total_cap
                        reason_override = (
                            f"Supplied — all {total_veh} vehicle(s) fully loaded; "
                            f"{deficit:.0f} L left at BMC | "
                            f"each vehicle empty = 0 L ≤ LQ {lq_val:.0f} L"
                        )
                    else:
                        min_cap = min((c for c in global_caps.values() if c > 0), default=10000.0)
                        do_not_supply = True
                        reason_override = (
                            f"Cannot supply: flow ({q:.0f} L) requires {min_cap:.0f} L vehicle but excess empty space exceeds {lq_val:.0f} L (15% rule)"
                        )

                total_disp_qty = 0.0
                if do_not_supply:
                    total_disp_qty = 0.0
                else:
                    if not reason_override:
                        if lq_val is not None:
                            reason_override = (
                                f"Supplied: partial vehicle empty {per_veh_empty:.0f} L "
                                f"≤ LQ {lq_val:.0f} L"
                            )
                        else:
                            reason_override = "Supplied"
                    total_disp_qty = total_cap if total_cap < q else q
                    
                    if key in subcluster_vehicle_pools:
                        for vc, count in optimal_veh.items():
                            if vc in subcluster_vehicle_pools[key]:
                                subcluster_vehicle_pools[key][vc] -= count

                # Apply to all routes in group
                for idx, r in enumerate(group):
                    r['margin_low'] = margin_low
                    r['margin_high'] = margin_high
                    r['min_flow_quantity'] = r['flow'] * (1.0 - margin_high / 100.0)
                    r['max_flow_quantity'] = r['flow'] * (1.0 + margin_low / 100.0)
                    
                    if idx == 0:
                        if do_not_supply:
                            global_caps = vehicle_limits_map.get('global_caps', {})
                            for vc in global_caps.keys():
                                r[f'vehicles_{vc}'] = 0
                            r['total_vehicles'] = 0
                            r['total_vehicle_capacity'] = 0.0
                            r['excess_vehicle_capacity'] = 0.0
                            r['per_vehicle_empty'] = 0.0
                            r['vehicle_reason'] = reason_override
                            r['dispatch_quantity'] = 0.0
                        else:
                            global_caps = vehicle_limits_map.get('global_caps', {})
                            for vc in global_caps.keys():
                                r[f'vehicles_{vc}'] = optimal_veh.get(vc, 0)
                            r['total_vehicles'] = total_veh
                            r['total_vehicle_capacity'] = total_cap
                            r['excess_vehicle_capacity'] = round(excess_qty, 2)
                            r['per_vehicle_empty'] = round(per_veh_empty, 2)
                            r['vehicle_reason'] = reason_override
                            r['dispatch_quantity'] = (r['flow'] / q) * total_disp_qty if q > 0 else 0.0
                    else:
                        global_caps = vehicle_limits_map.get('global_caps', {})
                        for vc in global_caps.keys():
                            r[f'vehicles_{vc}'] = 0
                        r['total_vehicles'] = 0
                        r['total_vehicle_capacity'] = 0.0
                        r['excess_vehicle_capacity'] = 0.0
                        r['per_vehicle_empty'] = 0.0
                        r['vehicle_reason'] = "Handled by Primary Row"
                        r['dispatch_quantity'] = (r['flow'] / q) * total_disp_qty if q > 0 else 0.0
            for r in res.get('routes', []):
                if r.get('from_type') != 'hub':
                    global_caps = vehicle_limits_map.get('global_caps', {})
                    for vc in global_caps.keys():
                        r[f'vehicles_{vc}'] = 0
                    r['total_vehicles'] = 0
                    r['total_vehicle_capacity'] = 0.0
                    r['excess_vehicle_capacity'] = 0.0
                    r['per_vehicle_empty'] = 0.0
                    r['vehicle_reason'] = "N/A"
                    r['dispatch_quantity'] = 0.0
                    r['margin_low'] = 0.0
                    r['margin_high'] = 0.0
                    r['min_flow_quantity'] = r['flow']
                    r['max_flow_quantity'] = r['flow']

            # 1. Summary Sheet
            summary_data = [
                {'Metric': 'Solver Status', 'Value': res['status']},
                {'Metric': 'Net Daily Profit (₹)', 'Value': res['summary']['profit']},
                {'Metric': 'Total Revenue (₹)', 'Value': res['summary']['revenue']},
                {'Metric': 'Total Transport Cost (₹)', 'Value': res['summary']['transport_cost']},
                {'Metric': 'Total Processing Cost (₹)', 'Value': res['summary']['processing_cost']},
                {'Metric': 'Hub Processing Cost (₹)', 'Value': res['summary']['hub_processing_cost']},
                {'Metric': 'Plant Processing Cost (₹)', 'Value': res['summary']['plant_processing_cost']},
                {'Metric': 'Total Nodes', 'Value': len(nodes)}
            ]
            
            # Calculate total flows by milk type
            flow_by_type = {}
            for r in res.get('routes', []):
                if r.get('from_type') == 'hub':
                    m = r.get('product_type', 'Unknown')
                    flow_by_type[m] = flow_by_type.get(m, 0.0) + r.get('flow', 0.0)
                    
            for m, total in flow_by_type.items():
                summary_data.append({'Metric': f'Total Flow - {m} (L)', 'Value': round(total, 2)})
                
            summary_data.append({'Metric': 'Processing Time (seconds)', 'Value': round(time.time() - start_time, 2)})
            
            df_summary = pd.DataFrame(summary_data)
            
            # 2. Nodes Sheet (Long Format)
            nodes_data = []
            
            inflow_lookup = {}
            outflow_lookup = {}
            for r in res.get('routes', []):
                fid = r['from_id']
                tid = r['to_id']
                ptype = r['product_type']
                flow_val = r['flow']
                
                inflow_lookup[(tid, ptype)] = inflow_lookup.get((tid, ptype), 0.0) + flow_val
                outflow_lookup[(fid, ptype)] = outflow_lookup.get((fid, ptype), 0.0) + flow_val

            for n in nodes:
                nid = n['id']
                ntype = n['type']

                if ntype == 'hub':
                    for p in n.get('products', []):
                        comm = p['type']
                        limit = p.get('capacity', 0.0)
                        inflow = inflow_lookup.get((nid, comm), 0.0)
                        outflow = outflow_lookup.get((nid, comm), 0.0)
                        nodes_data.append({
                            'Node ID': nid,
                            'Name': n['name'],
                            'Type': ntype,
                            'Subtype': n.get('subtype', ''),
                            'Latitude': n['lat'],
                            'Longitude': n['lng'],
                            'Commodity': comm,
                            'Inflow Throughput': round(inflow, 2),
                            'Outflow Throughput': round(outflow, 2),
                            'Capacity / Supply Limit': limit,
                                                    })
                elif ntype == 'plant':
                    p_demands = n.get('demands', [])
                    if p_demands:
                        for d in p_demands:
                            comm = d['type']
                            limit = d.get('demand', 0.0)
                            inflow = inflow_lookup.get((nid, comm), 0.0)
                            nodes_data.append({
                                'Node ID': nid,
                                'Name': n['name'],
                                'Type': ntype,
                                'Subtype': n.get('subtype', ''),
                                'Latitude': n['lat'],
                                'Longitude': n['lng'],
                                'Commodity': comm,
                                'Inflow Throughput': round(inflow, 2),
                                'Outflow Throughput': round(inflow, 2),
                                'Capacity / Supply Limit': limit,
                                                            })
                    else:
                        for m in n.get('inflow_milks', []):
                            comm = m['type']
                            limit = m.get('capacity', 0.0)
                            inflow = inflow_lookup.get((nid, comm), 0.0)
                            outflow = outflow_lookup.get((nid, comm), 0.0)
                            nodes_data.append({
                                'Node ID': nid,
                                'Name': n['name'],
                                'Type': ntype,
                                'Subtype': n.get('subtype', ''),
                                'Latitude': n['lat'],
                                'Longitude': n['lng'],
                                'Commodity': comm,
                                'Inflow Throughput': round(inflow, 2),
                                'Outflow Throughput': round(outflow, 2),
                                'Capacity / Supply Limit': limit,
                                                            })
                        for p in n.get('products', []):
                            comm = p['type']
                            limit = p.get('yield', 0.0)
                            inflow = inflow_lookup.get((nid, comm), 0.0)
                            outflow = outflow_lookup.get((nid, comm), 0.0)
                            nodes_data.append({
                                'Node ID': nid,
                                'Name': n['name'],
                                'Type': ntype,
                                'Subtype': n.get('subtype', ''),
                                'Latitude': n['lat'],
                                'Longitude': n['lng'],
                                'Commodity': comm,
                                'Inflow Throughput': round(inflow, 2),
                                'Outflow Throughput': round(outflow, 2),
                                'Capacity / Supply Limit': limit,
                                                            })

            df_nodes = pd.DataFrame(nodes_data)
            
            # Calculate plant-level substitution values from the solver
            plant_subs = {}
            for nd in nodes_data:
                if nd['Type'] == 'plant':
                    pid = nd['Node ID']
                    if pid not in plant_subs:
                        plant_subs[pid] = {
                            'BM_to_FCM': res.get('trans_vars_solution', {}).get((pid, 'BM_to_FCM'), 0.0),
                            'FCM_to_MM': res.get('trans_vars_solution', {}).get((pid, 'FCM_to_MM'), 0.0),
                            'BM_inflow': 0.0,
                            'FCM_inflow': 0.0
                        }
                    
                    comm = str(nd['Commodity']).lower()
                    inflow = nd['Inflow Throughput']
                    if 'buffalo' in comm or 'bm' in comm:
                        plant_subs[pid]['BM_inflow'] += inflow
                    elif 'fcm' in comm:
                        plant_subs[pid]['FCM_inflow'] += inflow
            
            # Since BM_to_FCM might cascade further to MM, we calculate how much BM went to MM
            # The amount of BM that goes to MM is any FCM_to_MM that exceeds physical FCM inflow.
            # But the plant pools them.
            for pid, subs in plant_subs.items():
                bm_to_fcm = subs['BM_to_FCM']
                fcm_to_mm = subs['FCM_to_MM']
                physical_fcm = subs['FCM_inflow']
                
                fcm_cascaded = min(fcm_to_mm, physical_fcm)
                bm_cascaded_to_mm = max(0, fcm_to_mm - physical_fcm)
                bm_cascaded_to_fcm = max(0, bm_to_fcm - bm_cascaded_to_mm)
                
                subs['BM_to_FCM_actual'] = bm_cascaded_to_fcm
                subs['BM_to_MM_actual'] = bm_cascaded_to_mm
                subs['FCM_to_MM_actual'] = fcm_cascaded
            
            # 3. Routes Sheet (Active and Unused/Candidate Routes)
            routes_data = []
            
            # Map of active routes to easily find flow details and avoid duplicates
            active_keys = {(r['from_id'], r['to_id'], r['product_type']) for r in res.get('routes', [])}
            
            # Populate active routes
            for r in res.get('routes', []):
                from_node = next((n for n in nodes if n['id'] == r['from_id']), {'name': 'Unknown'})
                to_node = next((n for n in nodes if n['id'] == r['to_id']), {'name': 'Unknown'})
                
                supplier = bmc_to_supplier.get(r['from_id'], '') if r['from_type'] == 'hub' else ''
                bmc_info = vehicle_limits_map.get(supplier, {}) if r['from_type'] == 'hub' else {}
                lq = bmc_info.get('leave_quantity', 0.0) if isinstance(bmc_info, dict) else 0.0
                cluster = bmc_info.get('cluster', '') if isinstance(bmc_info, dict) else ''
                subcluster = bmc_info.get('subcluster', '') if isinstance(bmc_info, dict) else ''
                strategy = bmc_info.get('strategy', '') if isinstance(bmc_info, dict) else ''
                
                ptype = r['product_type']
                ptype_lower = str(ptype).lower()
                
                # Default primary row values
                primary_flow = r['flow']
                extra_flows = [] # List of tuples: (flow, product_type, reason)
                # Append primary row
                routes_data.append({
                    'Route ID': r['id'],
                    'From Node ID': r['from_id'],
                    'From Name': from_node.get('name', 'Unknown'),
                    'From Type': r['from_type'],
                    'From Latitude': from_node.get('lat'),
                    'From Longitude': from_node.get('lng'),
                    'To Node ID': r['to_id'],
                    'To Name': to_node.get('name', 'Unknown'),
                    'To Type': r['to_type'],
                    'To Latitude': to_node.get('lat'),
                    'To Longitude': to_node.get('lng'),
                    'Product / Milk Type': ptype,
                    'Flow': max(0.0, round(primary_flow, 2)),
                    'Dispatch Quantity': max(0.0, round(primary_flow * (r.get('dispatch_quantity', 0.0) / r['flow']) if r.get('flow', 0) > 0 else 0.0, 2)),
                    'Unit': r['unit'],
                    'Distance (km)': r['distance'],
                    'Transport Cost (₹)': r['cost'],
                    'Status': 'ACTIVE',
                    'Reason': 'Converted Flow' if '_extra_' in r['id'] else 'Optimal Flow',
                    'Detailed Reason': 'Substituted internally at plant.' if '_extra_' in r['id'] else (f'Optimized for minimal transport cost (Dist: {r["distance"]} km) and even percentage balancing.' if r['from_type'] == 'hub' else f'Optimized to fulfill {ptype} market demand (Dist: {r["distance"]} km).'),
                    **{f'{vc} Vehicles': r.get(f'vehicles_{vc}', 0) for vc in vehicle_limits_map.get('global_caps', {}).keys()},
                    'Total Vehicles': r.get('vehicles_total', 0) if 'vehicles_total' in r else r.get('total_vehicles', 0),
                    'Total Vehicle Capacity (L)': r.get('capacity_total', 0) if 'capacity_total' in r else r.get('total_vehicle_capacity', 0),
                    'Excess Vehicle Capacity (L)': r.get('capacity_excess', 0) if 'capacity_excess' in r else r.get('excess_vehicle_capacity', 0),
                    'VehicleReason': r.get('vehicle_reason', 'Optimal'),
                    'SupplierCluster': cluster,
                    'SupplierSubCluster': subcluster,
                    'Strategy': strategy,
                    'FlowLowMarginPercentage': bmc_info.get('margin_low', 0.0) if isinstance(bmc_info, dict) else 5.0,
                    'FlowHighMarginPercentage': bmc_info.get('margin_high', 0.0) if isinstance(bmc_info, dict) else 5.0,
                    'MinimumFlowQuantity': bmc_info.get('min_flow', 0.0) if isinstance(bmc_info, dict) else 0.0,
                    'MaximumFlowQuantity': bmc_info.get('max_flow', 0.0) if isinstance(bmc_info, dict) else 0.0,
                    'Mapping Exists': 'Yes' if (str(r['from_id']), str(r['to_id']), str(r.get('target_product_type', ptype))) in valid_route_tuples else 'No'
                })

            # Pre-calculate lookup dictionaries for node capacities and demands
            hub_capacities = {}
            for h in hubs:
                h_prods = h.get('products', [])
                if not h_prods and 'capacity' in h:
                    h_prods = [{'type': 'Cow Milk', 'capacity': h['capacity']}]
                for p in h_prods:
                    if 'type' in p:
                        hub_capacities[(h['id'], p['type'])] = p.get('capacity', 0.0)

            plant_capacities = {}
            for p in plants:
                p_demands = p.get('demands', [])
                if p_demands:
                    for d in p_demands:
                        plant_capacities[(p['id'], d['type'])] = d.get('demand', 0.0)
                else:
                    inflow_milks = p.get('inflow_milks', [])
                    if not inflow_milks:
                        inflow_milks = [
                            {'type': 'Cow Milk', 'capacity': p.get('capacity', 10000)},
                            {'type': 'Buffalo Milk', 'capacity': p.get('capacity', 10000)}
                        ]
                    for m in inflow_milks:
                        if 'type' in m:
                            plant_capacities[(p['id'], m['type'])] = m.get('capacity', 0.0)



            # 3.1. Generate Unused Routes: Hub -> Plant
            for h in hubs:
                h_prods = h.get('products', [])
                if not h_prods and 'capacity' in h:
                    h_prods = [{'type': 'Cow Milk', 'capacity': h['capacity']}]
                for p in h_prods:
                    m = p.get('type')
                    if not m:
                        continue
                    for plant in plants:
                        p_demands = plant.get('demands', [])
                        plant_demand_types = {d['type'] for d in p_demands if 'type' in d}
                        
                        inflow_milks = plant.get('inflow_milks', [])
                        if not inflow_milks:
                            inflow_milks = [{'type': 'Cow Milk'}, {'type': 'Buffalo Milk'}]
                        plant_milk_types = {im['type'] for im in inflow_milks if 'type' in im}
                        
                        if m in plant_milk_types or m in plant_demand_types:
                            key = (h['id'], plant['id'], m)
                            if key not in active_keys:
                                total_h_out = outflow_lookup.get((h['id'], m), 0.0)
                                total_p_in = inflow_lookup.get((plant['id'], m), 0.0)
                                plant_cap = plant_capacities.get((plant['id'], m), 0.0)
                                hub_cap = hub_capacities.get((h['id'], m), 0.0)
                                
                                pair_key = tuple(sorted([h['id'], plant['id']]))
                                dist = _distance_cache.get(pair_key)
                                if dist is None:
                                    dist = 9999.0
                                    
                                if dist > MAX_DISTANCE_LIMIT:
                                    reason = f"Distance exceeds {int(MAX_DISTANCE_LIMIT)} km"
                                elif hub_cap <= 0:
                                    reason = "BMC capacity is zero"
                                elif plant_cap <= 0:
                                    reason = "Plant capacity/demand for this milk type is zero"
                                elif abs(total_h_out - hub_cap) < 0.1:
                                    reason = "BMC capacity is fully utilized by other Plants"
                                elif abs(total_p_in - plant_cap) < 0.1:
                                    reason = "Plant capacity/demand is fully utilized by other BMCs"
                                elif (str(h['id']), str(plant['id']), str(m)) not in valid_route_tuples:
                                    reason = "Mapping not exists."
                                else:
                                    reason = "Not selected by optimizer (sub-optimal route)."
                                routes_data.append({
                                    'Route ID': f"route_{h['id']}_{plant['id']}_{m.replace(' ', '_')}",
                                    'From Node ID': h['id'],
                                    'From Name': h.get('name', 'Unknown'),
                                    'From Type': 'hub',
                                    'From Latitude': h.get('lat'),
                                    'From Longitude': h.get('lng'),
                                    'To Node ID': plant['id'],
                                    'To Name': plant.get('name', 'Unknown'),
                                    'To Type': 'plant',
                                    'To Latitude': plant.get('lat'),
                                    'To Longitude': plant.get('lng'),
                                    'Product / Milk Type': m,
                                    'Flow': 0.0,
                                    'Unit': 'L',
                                    'Distance (km)': round(dist, 2),
                                    'Transport Cost (₹)': 0.0,
                                    'Status': 'UNUSED',
                                    'Reason': reason,
                                    'Detailed Reason': reason,
                                    **{f'{vc} Vehicles': 0 for vc in vehicle_limits_map.get('global_caps', {}).keys()},
                                    'Total Vehicles': 0,
                                    'Total Vehicle Capacity (L)': 0,
                                    'Excess Vehicle Capacity (L)': 0,
                                    'VehicleReason': 'Unused Route',
                                                                        'SupplierCluster': vehicle_limits_map.get(bmc_to_supplier.get(h['id'], ''), {}).get('cluster', '') if isinstance(vehicle_limits_map.get(bmc_to_supplier.get(h['id'], '')), dict) else '',
                                    'SupplierSubCluster': vehicle_limits_map.get(bmc_to_supplier.get(h['id'], ''), {}).get('subcluster', '') if isinstance(vehicle_limits_map.get(bmc_to_supplier.get(h['id'], '')), dict) else '',
                                    'Strategy': vehicle_limits_map.get(bmc_to_supplier.get(h['id'], ''), {}).get('strategy', '') if isinstance(vehicle_limits_map.get(bmc_to_supplier.get(h['id'], '')), dict) else '',
                                    'FlowLowMarginPercentage': vehicle_limits_map.get(bmc_to_supplier.get(h['id'], ''), {}).get('margin_low', 0.0) if isinstance(vehicle_limits_map.get(bmc_to_supplier.get(h['id'], '')), dict) else 5.0,
                                    'FlowHighMarginPercentage': vehicle_limits_map.get(bmc_to_supplier.get(h['id'], ''), {}).get('margin_high', 0.0) if isinstance(vehicle_limits_map.get(bmc_to_supplier.get(h['id'], '')), dict) else 5.0,
                                    'MinimumFlowQuantity': 0.0,
                                    'MaximumFlowQuantity': 0.0,
                                    'Mapping Exists': 'Yes' if (str(h['id']), str(plant['id']), str(m)) in valid_route_tuples else 'No'
                                })


            df_routes = pd.DataFrame(routes_data)
            
            # --- 1. Plant Consumption Report ---
            df_hub_to_plant = df_routes[
                (df_routes['Status'] == 'ACTIVE') & 
                (df_routes['From Type'] == 'hub') & 
                (df_routes['To Type'] == 'plant')
            ]
            
            unique_comms_set = set(df_nodes['Commodity'].unique()) if not df_nodes.empty else set()
            for _, r in df_hub_to_plant.iterrows():
                unique_comms_set.add(r['Product / Milk Type'])
            unique_commodities = sorted(list(unique_comms_set))
            # Group unique_commodities into base and conversions
            base_comms = [c for c in unique_commodities if ' to ' not in c]
            conv_comms = [c for c in unique_commodities if ' to ' in c]
            
            target_to_convs = {}
            for c in conv_comms:
                target = c.split(' to ')[1].strip()
                if target not in target_to_convs:
                    target_to_convs[target] = []
                target_to_convs[target].append(c)
                
            for target in target_to_convs:
                if target not in base_comms:
                    base_comms.append(target)
            base_comms.sort()
            
            plant_report_rows = []
            plant_received_report_rows = []
            for p in plants:
                row_dict = {'Plant ID': p['id'], 'Plant Name': p.get('name', p['id'])}
                row_received_dict = {'Plant ID': p['id'], 'Plant Name': p.get('name', p['id'])}
                plant_supplies = {}
                plant_received = {}
                for _, r in df_hub_to_plant.iterrows():
                    if r['To Node ID'] == p['id']:
                        c = r['Product / Milk Type']
                        plant_supplies[c] = plant_supplies.get(c, 0.0) + r['Flow']
                        plant_received[c] = plant_received.get(c, 0.0) + r.get('Dispatch Quantity', 0.0)
                        
                p_demands = p.get('demands', [])
                plant_demands = {}
                if p_demands:
                    for d in p_demands:
                        plant_demands[d['type']] = d.get('demand', 0.0)
                else:
                    inflow_milks = p.get('inflow_milks', [])
                    if not inflow_milks:
                        inflow_milks = [{'type': 'Cow Milk', 'capacity': p.get('capacity', 10000)},
                                        {'type': 'Buffalo Milk', 'capacity': p.get('capacity', 10000)}]
                    for m in inflow_milks:
                        if 'type' in m:
                            plant_demands[m['type']] = m.get('capacity', 0.0)
                
                for B in base_comms:
                    demand = plant_demands.get(B, 0.0)
                    supply = plant_supplies.get(B, 0.0)
                    received = plant_received.get(B, 0.0)
                    
                    if demand == 0:
                        pct = ""
                        received_pct = ""
                    else:
                        pct = round((supply / demand * 100.0), 2)
                        received_pct = round((received / demand * 100.0), 2)
                    
                    row_dict[f'{B} {{Demand}}'] = demand
                    row_dict[f'{B} {{Supply}}'] = supply
                    
                    row_received_dict[f'{B} {{Demand}}'] = demand
                    row_received_dict[f'{B} {{Received}}'] = received
                    
                    has_conv = B in target_to_convs
                    if not has_conv:
                        row_dict[f'{B} {{Received Percentage}}'] = pct
                        row_received_dict[f'{B} {{Received Percentage}}'] = received_pct
                    
                    final_supply = supply
                    final_received = received
                    if has_conv:
                        for C in target_to_convs[B]:
                            c_supply = plant_supplies.get(C, 0.0)
                            c_received = plant_received.get(C, 0.0)
                            row_dict[f'{C} {{Supply}}'] = c_supply
                            row_received_dict[f'{C} {{Received}}'] = c_received
                            final_supply += c_supply
                            final_received += c_received
                            
                        row_dict[f'Final {B} Supply'] = round(final_supply, 2)
                        row_received_dict[f'Final {B} Received'] = round(final_received, 2)
                        
                        final_pct = round((final_supply / demand * 100.0), 2) if demand > 0 else ""
                        final_received_pct = round((final_received / demand * 100.0), 2) if demand > 0 else ""
                        row_dict[f'Final {B} Percentage'] = final_pct
                        row_received_dict[f'Final {B} Percentage'] = final_received_pct
                        
                plant_report_rows.append(row_dict)
                plant_received_report_rows.append(row_received_dict)
                
            df_plant_report = pd.DataFrame(plant_report_rows)
            if df_plant_report.empty:
                df_plant_report = pd.DataFrame(columns=['Plant ID', 'Plant Name'])
                
            df_plant_received_report = pd.DataFrame(plant_received_report_rows)
            if df_plant_received_report.empty:
                df_plant_received_report = pd.DataFrame(columns=['Plant ID', 'Plant Name'])
            
            # --- 2. BMC Supply Report ---
            bmc_base_comms = [c for c in unique_commodities if ' to ' not in c]
            source_to_convs = {}
            for c in conv_comms:
                src = c.split(' to ')[0].strip()
                if src not in source_to_convs:
                    source_to_convs[src] = []
                source_to_convs[src].append(c)
                
            for src in source_to_convs:
                if src not in bmc_base_comms:
                    bmc_base_comms.append(src)
            bmc_base_comms.sort()
            
            hub_report_rows = []
            hub_dispatch_report_rows = []
            for h in hubs:
                supplier = bmc_to_supplier.get(h['id'], '')
                row_dict = {
                    'BMC ID': h['id'],
                    'BMC Name': h.get('name', h['id']),
                                    }
                row_dispatch_dict = {
                    'BMC ID': h['id'],
                    'BMC Name': h.get('name', h['id']),
                                    }
                
                hub_supplies = {}
                hub_dispatch = {}
                for _, r in df_hub_to_plant.iterrows():
                    if r['From Node ID'] == h['id']:
                        c = r['Product / Milk Type']
                        hub_supplies[c] = hub_supplies.get(c, 0.0) + r['Flow']
                        hub_dispatch[c] = hub_dispatch.get(c, 0.0) + r.get('Dispatch Quantity', 0.0)
                        
                h_prods = h.get('products', [])
                hub_stocks = {}
                if not h_prods and 'capacity' in h:
                    h_prods = [{'type': 'Cow Milk', 'capacity': h['capacity']}]
                for p_item in h_prods:
                    if p_item.get('type'):
                        hub_stocks[p_item['type']] = p_item.get('capacity', 0.0)
                        
                for B in bmc_base_comms:
                    stock = hub_stocks.get(B, 0.0)
                    supply = hub_supplies.get(B, 0.0)
                    dispatch = hub_dispatch.get(B, 0.0)
                    if stock == 0:
                        pct = ""
                        dispatch_pct = ""
                    else:
                        pct = round((supply / stock * 100.0), 2)
                        dispatch_pct = round((dispatch / stock * 100.0), 2)
                        
                    row_dict[f'{B} {{Stock}}'] = stock
                    row_dict[f'{B} {{Supply}}'] = supply
                    
                    row_dispatch_dict[f'{B} {{Stock}}'] = stock
                    row_dispatch_dict[f'{B} {{Dispatch}}'] = dispatch
                    
                    has_conv = B in source_to_convs
                    if not has_conv:
                        row_dict[f'{B} {{Supply Percentage}}'] = pct
                        row_dispatch_dict[f'{B} {{Dispatch Percentage}}'] = dispatch_pct
                        
                    final_supply = supply
                    final_dispatch = dispatch
                    if has_conv:
                        for C in source_to_convs[B]:
                            c_supply = hub_supplies.get(C, 0.0)
                            c_dispatch = hub_dispatch.get(C, 0.0)
                            row_dict[f'{C} {{Supply}}'] = c_supply
                            row_dispatch_dict[f'{C} {{Dispatch}}'] = c_dispatch
                            final_supply += c_supply
                            final_dispatch += c_dispatch
                            
                        row_dict[f'Final {B} Supply'] = round(final_supply, 2)
                        row_dispatch_dict[f'Final {B} Dispatch'] = round(final_dispatch, 2)
                        
                        final_pct = round((final_supply / stock * 100.0), 2) if stock > 0 else ""
                        final_dispatch_pct = round((final_dispatch / stock * 100.0), 2) if stock > 0 else ""
                        row_dict[f'Final {B} Percentage'] = final_pct
                        row_dispatch_dict[f'Final {B} Percentage'] = final_dispatch_pct
                hub_report_rows.append(row_dict)
                hub_dispatch_report_rows.append(row_dispatch_dict)
                
            df_hub_report = pd.DataFrame(hub_report_rows)
            if df_hub_report.empty:
                df_hub_report = pd.DataFrame(columns=['BMC ID', 'BMC Name'])
                
            df_hub_dispatch_report = pd.DataFrame(hub_dispatch_report_rows)
            if df_hub_dispatch_report.empty:
                df_hub_dispatch_report = pd.DataFrame(columns=['BMC ID', 'BMC Name'])
                
            # --- 3. Hub To Plant (optimal flow Hub -> Plant) ---
            # df_hub_to_plant is already generated above
            
            # --- 4. BMC Wise Allocation Matrix ---
            plant_names_list = []
            seen_plants = set()
            for p in plants:
                name = p.get('name', p['id'])
                if name not in seen_plants:
                    seen_plants.add(name)
                    plant_names_list.append(name)
            
            flow_map = {}
            dispatch_map = {}
            for _, r in df_hub_to_plant.iterrows():
                key = (r['From Node ID'], r['To Node ID'], r['Product / Milk Type'])
                flow_map[key] = flow_map.get(key, 0.0) + r['Flow']
                dispatch_map[key] = dispatch_map.get(key, 0.0) + r.get('Dispatch Quantity', 0.0)
                    
            bmc_allocation_rows = []
            bmc_dispatch_rows = []
            for h in hubs:
                h_prods = h.get('products', [])
                if not h_prods and 'capacity' in h:
                    h_prods = [{'type': 'Cow Milk', 'capacity': h['capacity']}]
                
                hub_commodities = set()
                for p_item in h_prods:
                    if p_item.get('type'):
                        hub_commodities.add(p_item['type'])
                for (from_id, to_id, prod_type) in flow_map:
                    if from_id == h['id']:
                        hub_commodities.add(prod_type)
                        
                for m in sorted(list(hub_commodities)):
                    total_flow = sum(flow_map.get((h['id'], p['id'], m), 0.0) for p in plants)
                    total_dispatch = sum(dispatch_map.get((h['id'], p['id'], m), 0.0) for p in plants)
                    
                    if round(total_flow, 2) > 0:
                        supplier = bmc_to_supplier.get(h['id'], '')
                        bmc_info = vehicle_limits_map.get(supplier, {})
                        lq = bmc_info.get('leave_quantity', 0.0) if isinstance(bmc_info, dict) else 0.0
                        row_dict = {
                            'BMC': h.get('name', h['id']),
                            'Product': m,
                            'Quantity': round(total_flow, 2)
                        }
                        for p in plants:
                            p_name = p.get('name', p['id'])
                            flow_val = flow_map.get((h['id'], p['id'], m), 0.0)
                            row_dict[p_name] = round(flow_val, 2)
                        bmc_allocation_rows.append(row_dict)
                        
                    if round(total_dispatch, 2) > 0:
                        row_dispatch_dict = {
                            'BMC': h.get('name', h['id']),
                            'Product': m,
                            'Quantity': round(total_dispatch, 2)
                        }
                        for p in plants:
                            p_name = p.get('name', p['id'])
                            disp_val = dispatch_map.get((h['id'], p['id'], m), 0.0)
                            row_dispatch_dict[p_name] = round(disp_val, 2)
                        bmc_dispatch_rows.append(row_dispatch_dict)
                    
            df_bmc_wise_alloc = pd.DataFrame(bmc_allocation_rows)
            if not df_bmc_wise_alloc.empty:
                cols = ['BMC', 'Product', 'Quantity'] + [p_name for p_name in plant_names_list if p_name in df_bmc_wise_alloc.columns]
                df_bmc_wise_alloc = df_bmc_wise_alloc.reindex(columns=cols)
            else:
                df_bmc_wise_alloc = pd.DataFrame(columns=['BMC', 'Product', 'Quantity'] + plant_names_list)
                
            df_bmc_wise_dispatch = pd.DataFrame(bmc_dispatch_rows)
            if not df_bmc_wise_dispatch.empty:
                cols = ['BMC', 'Product', 'Quantity'] + [p_name for p_name in plant_names_list if p_name in df_bmc_wise_dispatch.columns]
                df_bmc_wise_dispatch = df_bmc_wise_dispatch.reindex(columns=cols)
            else:
                df_bmc_wise_dispatch = pd.DataFrame(columns=['BMC', 'Product', 'Quantity'] + plant_names_list)
 
            # --- 5. Plant Wise Allocation Matrix ---
            bmc_names_list = []
            seen_bmcs = set()
            for h in hubs:
                name = h.get('name', h['id'])
                if name not in seen_bmcs:
                    seen_bmcs.add(name)
                    bmc_names_list.append(name)
                    
            plant_allocation_rows = []
            plant_dispatch_rows = []
            for p in plants:
                p_commodities = {}
                p_demands = p.get('demands', [])
                if p_demands:
                    for d in p_demands:
                        p_commodities[d['type']] = d.get('demand', 0.0)
                else:
                    inflow_milks = p.get('inflow_milks', [])
                    if not inflow_milks:
                        inflow_milks = [
                            {'type': 'Cow Milk', 'capacity': p.get('capacity', 10000)},
                            {'type': 'Buffalo Milk', 'capacity': p.get('capacity', 10000)}
                        ]
                    for m in inflow_milks:
                        if 'type' in m:
                            p_commodities[m['type']] = m.get('capacity', 0.0)
                            
                for (from_id, to_id, prod_type) in flow_map:
                    if to_id == p['id'] and prod_type not in p_commodities:
                        p_commodities[prod_type] = 0.0
                        
                for m, req_qty in sorted(p_commodities.items()):
                    fulfilled_qty = sum(flow_map.get((h['id'], p['id'], m), 0.0) for h in hubs)
                    dispatch_qty = sum(dispatch_map.get((h['id'], p['id'], m), 0.0) for h in hubs)
                    
                    if req_qty == 0:
                        pct = ""
                        dispatch_pct = ""
                    else:
                        pct = round((fulfilled_qty / req_qty * 100.0), 2)
                        dispatch_pct = round((dispatch_qty / req_qty * 100.0), 2)
                        
                    row_dict = {
                        'Plant': p.get('name', p['id']),
                        'Product': m,
                        'Required Quantity': round(req_qty, 2),
                        'Fullfilled Quantity': round(fulfilled_qty, 2),
                        'Fullfilled Percentage': pct
                    }
                    for h in hubs:
                        h_name = h.get('name', h['id'])
                        flow_val = flow_map.get((h['id'], p['id'], m), 0.0)
                        row_dict[h_name] = round(flow_val, 2)
                    plant_allocation_rows.append(row_dict)
                    
                    row_dispatch_dict = {
                        'Plant': p.get('name', p['id']),
                        'Product': m,
                        'Required Quantity': round(req_qty, 2),
                        'Fullfilled Quantity': round(dispatch_qty, 2),
                        'Fullfilled Percentage': dispatch_pct
                    }
                    for h in hubs:
                        h_name = h.get('name', h['id'])
                        disp_val = dispatch_map.get((h['id'], p['id'], m), 0.0)
                        row_dispatch_dict[h_name] = round(disp_val, 2)
                    plant_dispatch_rows.append(row_dispatch_dict)
                    
            df_plant_wise_alloc = pd.DataFrame(plant_allocation_rows)
            if not df_plant_wise_alloc.empty:
                cols = ['Plant', 'Product', 'Required Quantity', 'Fullfilled Quantity', 'Fullfilled Percentage'] + [h_name for h_name in bmc_names_list if h_name in df_plant_wise_alloc.columns]
                df_plant_wise_alloc = df_plant_wise_alloc.reindex(columns=cols)
            else:
                df_plant_wise_alloc = pd.DataFrame(columns=['Plant', 'Product', 'Required Quantity', 'Fullfilled Quantity', 'Fullfilled Percentage'] + bmc_names_list)
                
            df_plant_wise_dispatch = pd.DataFrame(plant_dispatch_rows)
            if not df_plant_wise_dispatch.empty:
                cols = ['Plant', 'Product', 'Required Quantity', 'Fullfilled Quantity', 'Fullfilled Percentage'] + [h_name for h_name in bmc_names_list if h_name in df_plant_wise_dispatch.columns]
                df_plant_wise_dispatch = df_plant_wise_dispatch.reindex(columns=cols)
                
                # Update quantities for converted milk
                mask = df_plant_wise_dispatch['Product'].astype(str).str.contains('to FCM|to MM', na=False)
                converted_rows = df_plant_wise_dispatch[mask]
                
                for _, row in converted_rows.iterrows():
                    plant = row['Plant']
                    product = row['Product']
                    if ' to ' in product:
                        converted_milk = product.split(' to ')[1].strip()
                    else:
                        continue
                        
                    target_mask = (df_plant_wise_dispatch['Plant'].astype(str).str.strip().str.lower() == str(plant).strip().lower()) & (df_plant_wise_dispatch['Product'].astype(str).str.strip().str.lower() == str(converted_milk).strip().lower())
                    if target_mask.any():
                        target_idx = df_plant_wise_dispatch.index[target_mask][0]
                        
                        converted_fulfilled_qty = pd.to_numeric(row['Fullfilled Quantity'], errors='coerce')
                        converted_fulfilled_qty = converted_fulfilled_qty if pd.notna(converted_fulfilled_qty) else 0
                        
                        target_fulfilled_qty = pd.to_numeric(df_plant_wise_dispatch.at[target_idx, 'Fullfilled Quantity'], errors='coerce')
                        target_fulfilled_qty = target_fulfilled_qty if pd.notna(target_fulfilled_qty) else 0
                        
                        req_qty = pd.to_numeric(df_plant_wise_dispatch.at[target_idx, 'Required Quantity'], errors='coerce')
                        req_qty = req_qty if pd.notna(req_qty) else 0
                        
                        total_fulfilled = target_fulfilled_qty + converted_fulfilled_qty
                        
                        if req_qty > 0:
                            df_plant_wise_dispatch.at[target_idx, 'Fullfilled Percentage'] = round((total_fulfilled / req_qty) * 100, 2)
                        else:
                            df_plant_wise_dispatch.at[target_idx, 'Fullfilled Percentage'] = ""
            else:
                df_plant_wise_dispatch = pd.DataFrame(columns=['Plant', 'Product', 'Required Quantity', 'Fullfilled Quantity', 'Fullfilled Percentage'] + bmc_names_list)
            
            # Calculate KPIs
            try:
                total_flow = df_hub_to_plant['Flow'].sum() if not df_hub_to_plant.empty else 0.0
                total_vehicles = df_hub_to_plant['Total Vehicles'].sum() if not df_hub_to_plant.empty else 0
                total_driven_km = (df_hub_to_plant['Distance (km)'] * df_hub_to_plant['Total Vehicles']).sum() if not df_hub_to_plant.empty else 0.0

                trips_per_1m = (total_vehicles / total_flow) * 1000000 if total_flow > 0 else 0.0
                avg_l_per_trip = (total_flow / total_vehicles) if total_vehicles > 0 else 0.0
                km_per_1000l = (total_driven_km / total_flow) * 1000 if total_flow > 0 else 0.0

                df_kpi = pd.DataFrame([{
                    'Trips': int(total_vehicles),
                    'Trips/1M L': round(trips_per_1m, 2),
                    'Avg L/trip': round(avg_l_per_trip, 2),
                    'KM/1000 L': round(km_per_1000l, 2)
                }])
            except Exception as e:
                print("Error calculating KPIs:", e)
                df_kpi = pd.DataFrame([{'Trips': 0, 'Trips/1M L': 0, 'Avg L/trip': 0, 'KM/1000 L': 0}])

            # Save sheets
            
            # Create and save a dedicated BMC Vehicle Allocation sheet
            veh_cols = [f'{vc} Vehicles' for vc in vehicle_limits_map.get('global_caps', {}).keys()]
            df_veh_alloc = df_hub_to_plant[[
                'From Node ID', 'From Name', 'To Node ID', 'To Name', 
                'Product / Milk Type', 'Flow', 'Dispatch Quantity', 'Unit', 'Distance (km)', 'Transport Cost (₹)'
            ] + veh_cols + [
                'Total Vehicles', 'Total Vehicle Capacity (L)', 'Excess Vehicle Capacity (L)',
                'SupplierCluster', 'SupplierSubCluster', 'Strategy',
                'FlowLowMarginPercentage', 'FlowHighMarginPercentage',
                'MinimumFlowQuantity', 'MaximumFlowQuantity', 'VehicleReason'
            ]].copy()
            df_veh_alloc['Left Quantity'] = df_veh_alloc['Flow'] - df_veh_alloc['Dispatch Quantity']
            df_veh_alloc = df_veh_alloc.rename(columns={
                'From Node ID': 'BMC ID',
                'From Name': 'BMC Name',
                'To Node ID': 'Plant ID',
                'To Name': 'Plant Name',
                'Flow': 'Flow Quantity'
            })
            
            # --- Vehicle Wise Report ---
            v_codes = ['V07', 'V10', 'V12', 'V15', 'V20', 'V25', 'V30', 'V35']
            v_cols_present = [f"{v} Vehicles" for v in v_codes if f"{v} Vehicles" in df_veh_alloc.columns]
            
            max_dist = df_veh_alloc['Distance (km)'].max()
            if pd.isna(max_dist) or max_dist == 0:
                max_dist = 150
            max_dist_val = int(max_dist)
            bins = list(range(0, ((max_dist_val // 150) + 2) * 150, 150))
            labels = [f"{bins[i]}-{bins[i+1]}" for i in range(len(bins)-1)]
            df_veh_alloc_copy = df_veh_alloc.copy()
            df_veh_alloc_copy['Distance Range'] = pd.cut(df_veh_alloc_copy['Distance (km)'], bins=bins, labels=labels, right=False, include_lowest=True)
            
            vehicle_report_dist = df_veh_alloc_copy.groupby('Distance Range', observed=True)[v_cols_present].sum().T
            vehicle_report_dist.index = [c.replace(' Vehicles', '') for c in vehicle_report_dist.index]
            vehicle_report_dist.index.name = 'Vehicle Category'
            vehicle_report_dist.reset_index(inplace=True)
            
            dist_columns = [col for col in vehicle_report_dist.columns if col != 'Vehicle Category']
            for col in dist_columns:
                vehicle_report_dist[col] = pd.to_numeric(vehicle_report_dist[col], errors='coerce').fillna(0)
            vehicle_report_dist['Total'] = vehicle_report_dist[dist_columns].sum(axis=1)
            
            total_row_dist = pd.Series(index=vehicle_report_dist.columns, dtype='object')
            total_row_dist['Vehicle Category'] = 'Grand Total'
            for col in dist_columns + ['Total']:
                total_row_dist[col] = pd.to_numeric(vehicle_report_dist[col], errors='coerce').sum()
                
            # --- Total Supply Report ---
            df_total_supply_temp = df_hub_to_plant.copy()
            if not df_total_supply_temp.empty:
                df_total_supply_temp['Base Milk'] = df_total_supply_temp['Product / Milk Type'].apply(lambda x: x.split(' to ')[1].strip() if ' to ' in str(x) else str(x).strip())
                df_total_supply_temp = df_total_supply_temp.drop(columns=['Product / Milk Type', 'Route ID'], errors='ignore')
                
                group_cols = ['From Node ID', 'To Node ID', 'Base Milk']
                agg_dict = {}
                for col in df_total_supply_temp.columns:
                    if col in group_cols:
                        continue
                    if col in ['Flow', 'Dispatch Quantity', 'Transport Cost (₹)', 'Total Vehicles', 'Total Vehicle Capacity (L)', 'Excess Vehicle Capacity (L)'] or str(col).endswith(' Vehicles'):
                        agg_dict[col] = 'sum'
                    else:
                        agg_dict[col] = 'first'
                
                df_total_supply = df_total_supply_temp.groupby(group_cols, observed=True).agg(agg_dict).reset_index()
                
                # Reconstruct Route ID
                df_total_supply['Route ID'] = 'route_' + df_total_supply['From Node ID'].astype(str) + '_' + df_total_supply['To Node ID'].astype(str) + '_' + df_total_supply['Base Milk'].astype(str).str.replace(' ', '_')
                
                original_cols = list(df_hub_to_plant.columns)
                if 'Product / Milk Type' in original_cols:
                    idx = original_cols.index('Product / Milk Type')
                    new_cols = original_cols[:idx] + ['Base Milk'] + original_cols[idx+1:]
                else:
                    new_cols = original_cols + ['Base Milk']
                    
                df_total_supply = df_total_supply[[c for c in new_cols if c in df_total_supply.columns]]
                
                # Add Total Distance column
                if 'Dispatch Quantity' in df_total_supply.columns:
                    df_total_supply['Total Distance'] = df_total_supply.apply(lambda r: (r.get('Total Vehicles', 0) * r.get('Distance (km)', 0)) if r.get('Dispatch Quantity', 0) > 0 else 0.0, axis=1)
                else:
                    df_total_supply['Total Distance'] = df_total_supply.get('Total Vehicles', 0) * df_total_supply.get('Distance (km)', 0)
                    
                # Add prorated Dispatch Capacity columns
                global_caps = vehicle_limits_map.get('global_caps', {})
                for vc, cap in global_caps.items():
                    v_col = f'{vc} Vehicles'
                    new_col = f'{vc} Dispatch Capacity'
                    if v_col in df_total_supply.columns:
                        if 'Dispatch Quantity' in df_total_supply.columns:
                            df_total_supply[new_col] = df_total_supply.apply(
                                lambda r: round(r['Dispatch Quantity'] * (r[v_col] * cap) / r['Total Vehicle Capacity (L)'], 2) if r.get('Total Vehicle Capacity (L)', 0) > 0 else 0.0, 
                                axis=1
                            )
                        else:
                            df_total_supply[new_col] = df_total_supply.apply(
                                lambda r: round(r['Flow'] * (r[v_col] * cap) / r['Total Vehicle Capacity (L)'], 2) if r.get('Total Vehicle Capacity (L)', 0) > 0 else 0.0, 
                                axis=1
                            )
            else:
                cols = [c if c != 'Product / Milk Type' else 'Base Milk' for c in df_hub_to_plant.columns] + ['Total Distance']
                global_caps = vehicle_limits_map.get('global_caps', {})
                for vc in global_caps.keys():
                    cols.append(f'{vc} Dispatch Capacity')
                df_total_supply = pd.DataFrame(columns=cols)
            # --- Vehicle Wise Bifurcation Report ---
            veh_bifurcation_rows = []
            global_caps = vehicle_limits_map.get('global_caps', {})
            for vc, cap in global_caps.items():
                v_col = f'{vc} Vehicles'
                if v_col in df_total_supply.columns and df_total_supply[v_col].sum() > 0:
                    total_veh = df_total_supply[v_col].sum()
                    total_dist = (df_total_supply['Distance (km)'] * df_total_supply[v_col]).sum()
                    
                    disp_cap_col = f'{vc} Dispatch Capacity'
                    if disp_cap_col in df_total_supply.columns:
                        total_supply_v = df_total_supply[disp_cap_col].sum()
                    else:
                        supply_series = df_total_supply['Flow'] * (df_total_supply[v_col] * cap) / df_total_supply['Total Vehicle Capacity (L)'].replace({0: 1})
                        supply_series = supply_series.where(df_total_supply['Total Vehicle Capacity (L)'] > 0, 0)
                        total_supply_v = supply_series.sum()
                    
                    veh_bifurcation_rows.append({
                        'Vehicle Category': vc,
                        'Total Distance': total_dist,
                        'Dispatch Capacity': total_supply_v,
                        'TotalVehicle': total_veh
                    })
                    
            df_vehicle_bifurcation = pd.DataFrame(veh_bifurcation_rows)
            if not df_vehicle_bifurcation.empty:
                total_row_bif = {
                    'Vehicle Category': 'Grand Total',
                    'Total Distance': df_vehicle_bifurcation['Total Distance'].sum(),
                    'Dispatch Capacity': df_vehicle_bifurcation['Dispatch Capacity'].sum(),
                    'TotalVehicle': df_vehicle_bifurcation['TotalVehicle'].sum()
                }
                df_vehicle_bifurcation = pd.concat([df_vehicle_bifurcation, pd.DataFrame([total_row_bif])], ignore_index=True)
            else:
                df_vehicle_bifurcation = pd.DataFrame(columns=['Vehicle Category', 'Total Distance', 'Dispatch Capacity', 'TotalVehicle'])

            df_vehicle_wise_report = pd.concat([vehicle_report_dist, pd.DataFrame([total_row_dist])], ignore_index=True)
            
            # --- New Logic: Maximize Vehicle Utilization ---
            df_routes_optimized = df_routes.copy()
            
            # Merge 'Handled by Primary Row' directly into primary rows in Routes (Max Utilized) report
            if 'VehicleReason' in df_routes_optimized.columns:
                handled_mask = df_routes_optimized['VehicleReason'] == 'Handled by Primary Row'
                handled_indices = df_routes_optimized[handled_mask].index.tolist()
                
                for sec_idx in handled_indices:
                    from_id = df_routes_optimized.at[sec_idx, 'From Node ID']
                    to_id = df_routes_optimized.at[sec_idx, 'To Node ID']
                    sec_prod = str(df_routes_optimized.at[sec_idx, 'Product / Milk Type'])
                    target_comm = sec_prod.split(' to ')[-1].strip() if ' to ' in sec_prod else sec_prod.strip()
                    
                    # Locate the corresponding primary row candidate
                    candidates = df_routes_optimized[
                        (df_routes_optimized['From Node ID'] == from_id) &
                        (df_routes_optimized['To Node ID'] == to_id) &
                        (df_routes_optimized['VehicleReason'] != 'Handled by Primary Row') &
                        (~df_routes_optimized.index.isin(handled_indices))
                    ].index.tolist()
                    
                    pri_idx = None
                    for c_idx in candidates:
                        c_prod = str(df_routes_optimized.at[c_idx, 'Product / Milk Type'])
                        c_target = c_prod.split(' to ')[-1].strip() if ' to ' in c_prod else c_prod.strip()
                        if c_target == target_comm:
                            pri_idx = c_idx
                            break
                    if pri_idx is None and len(candidates) > 0:
                        pri_idx = candidates[0]
                        
                    if pri_idx is not None:
                        df_routes_optimized.at[pri_idx, 'Flow'] = round(df_routes_optimized.at[pri_idx, 'Flow'] + df_routes_optimized.at[sec_idx, 'Flow'], 2)
                        df_routes_optimized.at[pri_idx, 'Dispatch Quantity'] = round(df_routes_optimized.at[pri_idx, 'Dispatch Quantity'] + df_routes_optimized.at[sec_idx, 'Dispatch Quantity'], 2)
                
                # Drop secondary rows so two rows aren't created in Routes (Max Utilized)
                df_routes_optimized = df_routes_optimized[~handled_mask].reset_index(drop=True)
            
            # Add new tracking columns for visibility
            df_routes_optimized['Original Flow'] = df_routes_optimized['Flow']
            df_routes_optimized['Added Flow'] = 0.0
            df_routes_optimized['Original Dispatch Quantity'] = df_routes_optimized['Dispatch Quantity']
            df_routes_optimized['Added Dispatch Quantity'] = 0.0
            df_routes_optimized['Original Excess Vehicle Capacity (L)'] = df_routes_optimized['Excess Vehicle Capacity (L)']
            
            added_details = {} # { t_idx: [(added_amount, source_id, target_comm), ...] }
            
            # 1. Setup mappings from plant_bmc_mapping for fallback routes
            bmc_mappings = {}
            if 'plant_bmc_mapping' in locals():
                for p_code, bmc_code, comm in plant_bmc_mapping:
                    if bmc_code not in bmc_mappings:
                        bmc_mappings[bmc_code] = []
                    if (p_code, comm) not in bmc_mappings[bmc_code]:
                        bmc_mappings[bmc_code].append((p_code, comm))

            # 2. Iterate over all un-dispatched rows (Dispatch Quantity == 0)
            mask_undispatched = (df_routes_optimized['Dispatch Quantity'] == 0)
            undispatched_indices = df_routes_optimized[mask_undispatched].index.tolist()
            
            for u_idx in undispatched_indices:
                amount_to_send = df_routes_optimized.at[u_idx, 'Flow']
                if amount_to_send <= 0:
                    continue
                    
                h_id = df_routes_optimized.at[u_idx, 'From Node ID']
                p_id = df_routes_optimized.at[u_idx, 'To Node ID']
                m_type = str(df_routes_optimized.at[u_idx, 'Product / Milk Type']).strip()
                m_type_lower = m_type.lower()
                
                priorities = []
                allowed_comms = []
                if m_type_lower == "bm to fcm":
                    priorities.append((p_id, "FCM"))
                    priorities.append((p_id, "BM"))
                    allowed_comms = ["FCM", "BM"]
                elif m_type_lower == "bm to mm":
                    priorities.append((p_id, "MM"))
                    priorities.append((p_id, "BM"))
                    allowed_comms = ["MM", "BM"]
                elif m_type_lower == "fcm to mm":
                    priorities.append((p_id, "MM"))
                    priorities.append((p_id, "FCM"))
                    allowed_comms = ["MM", "FCM"]
                else:
                    priorities.append((p_id, m_type))
                    allowed_comms = [m_type.upper()]
                    
                for mapped_p, mapped_comm in bmc_mappings.get(h_id, []):
                    # Only allow mappings that respect the strict segregation rules
                    if mapped_comm.upper() in allowed_comms:
                        if (mapped_p, mapped_comm) not in priorities:
                            priorities.append((mapped_p, mapped_comm))
                        
                for target_p, target_comm in priorities:
                    if amount_to_send <= 0:
                        break
                        
                    # Find base vehicles from h_id to target_p with excess capacity
                    mask_veh = (
                        (df_routes_optimized['From Node ID'] == h_id) &
                        (df_routes_optimized['To Node ID'] == target_p) &
                        (df_routes_optimized['Excess Vehicle Capacity (L)'] > 0) &
                        (~df_routes_optimized['Route ID'].astype(str).str.contains('_extra_')) &
                        (df_routes_optimized['Status'] == 'ACTIVE')
                    )
                    veh_indices = df_routes_optimized[mask_veh].index.tolist()
                    
                    for v_idx in veh_indices:
                        if amount_to_send <= 0:
                            break
                            
                        # Strict Segregation Rule:
                        # Any milk being added to a vehicle MUST EXACTLY match that vehicle's base milk type.
                        # Since target_comm is already constrained by the allowed_comms logic above,
                        # enforcing this guarantees BM only goes on BM, CM on CM, FCM on FCM, and MM on MM.
                        base_v_m_type = str(df_routes_optimized.at[v_idx, 'Product / Milk Type']).strip().upper()
                        target_m_type = target_comm.upper()
                        
                        if base_v_m_type != target_m_type:
                            continue
                            
                        excess = df_routes_optimized.at[v_idx, 'Excess Vehicle Capacity (L)']
                        if excess <= 0:
                            continue
                            
                        added = min(amount_to_send, excess)
                        veh_route_id = str(df_routes_optimized.at[v_idx, 'Route ID'])
                        
                        t_idx = v_idx
                        df_routes_optimized.at[t_idx, 'Flow'] = round(df_routes_optimized.at[t_idx, 'Flow'] + added, 2)
                        df_routes_optimized.at[t_idx, 'Added Flow'] = round(df_routes_optimized.at[t_idx, 'Added Flow'] + added, 2)
                        df_routes_optimized.at[t_idx, 'Dispatch Quantity'] = round(df_routes_optimized.at[t_idx, 'Dispatch Quantity'] + added, 2)
                        df_routes_optimized.at[t_idx, 'Added Dispatch Quantity'] = round(df_routes_optimized.at[t_idx, 'Added Dispatch Quantity'] + added, 2)
                        
                        if t_idx not in added_details:
                            added_details[t_idx] = []
                        orig_m_type_raw = str(df_routes_optimized.at[u_idx, 'Product / Milk Type']).strip()
                        added_details[t_idx].append((added, str(df_routes_optimized.at[u_idx, 'Route ID']), target_comm, orig_m_type_raw))
                            
                        current_reason = str(df_routes_optimized.at[t_idx, 'Detailed Reason'])
                        orig_m_type = m_type.upper()
                        new_reason_part = f"Loaded {added}L of un-dispatched {orig_m_type} (as {target_comm}) from {df_routes_optimized.at[u_idx, 'Route ID']}"
                        if current_reason == "N/A" or pd.isna(df_routes_optimized.at[t_idx, 'Detailed Reason']) or current_reason == "nan":
                            df_routes_optimized.at[t_idx, 'Detailed Reason'] = new_reason_part
                        else:
                            df_routes_optimized.at[t_idx, 'Detailed Reason'] = current_reason + " | " + new_reason_part
                            
                        df_routes_optimized.at[v_idx, 'Excess Vehicle Capacity (L)'] = max(0.0, round(df_routes_optimized.at[v_idx, 'Excess Vehicle Capacity (L)'] - added, 2))
                        amount_to_send -= added
                        
                # Deduct sent amount from the undispatched row to avoid double counting Flow overall
                sent_amount = df_routes_optimized.at[u_idx, 'Flow'] - amount_to_send
                if sent_amount > 0:
                    df_routes_optimized.at[u_idx, 'Flow'] = round(amount_to_send, 2)
                    df_routes_optimized.at[u_idx, 'Detailed Reason'] = str(df_routes_optimized.at[u_idx, 'Detailed Reason']) + f" | {sent_amount}L sent along with other routes."
            
            # 3. Generate dynamic Added Flow columns based on added_details
            if added_details:
                max_added = max([len(v) for v in added_details.values()])
                for i in range(1, max_added + 1):
                    df_routes_optimized[f'Added Flow {i}'] = 0.0
                    df_routes_optimized[f'Added Flow Source {i}'] = "N/A"
                    df_routes_optimized[f'Added Flow Milk Type {i}'] = "N/A"
                    df_routes_optimized[f'Added Flow Original Milk Type {i}'] = "N/A"
                
                for t_idx, details in added_details.items():
                    for i, (amount, source, milk_type, orig_milk_type) in enumerate(details):
                        col_idx = i + 1
                        df_routes_optimized.at[t_idx, f'Added Flow {col_idx}'] = amount
                        df_routes_optimized.at[t_idx, f'Added Flow Source {col_idx}'] = source
                        df_routes_optimized.at[t_idx, f'Added Flow Milk Type {col_idx}'] = milk_type
                        df_routes_optimized.at[t_idx, f'Added Flow Original Milk Type {col_idx}'] = orig_milk_type
            
            # Add Target Milk column right after Product / Milk Type
            df_routes_optimized['Target Milk'] = df_routes_optimized['Product / Milk Type'].apply(
                lambda x: str(x).split(' to ')[1].strip() if ' to ' in str(x).lower() else str(x).strip()
            )
            if 'Product / Milk Type' in df_routes_optimized.columns:
                cols_opt = list(df_routes_optimized.columns)
                cols_opt.remove('Target Milk')
                idx = cols_opt.index('Product / Milk Type') + 1
                cols_opt = cols_opt[:idx] + ['Target Milk'] + cols_opt[idx:]
                df_routes_optimized = df_routes_optimized[cols_opt]
            
            # --- Generate BMC Wise Dispatch (Max Utilized) ---
            dispatch_map_opt = {}
            converted_map_opt = {}
            converted_types = set()
            
            df_opt_active = df_routes_optimized[df_routes_optimized['Status'] == 'ACTIVE']
            for _, r in df_opt_active.iterrows():
                from_id = r['From Node ID']
                to_id = r['To Node ID']
                
                # Base flow of the vehicle
                raw_base_m_type = str(r['Product / Milk Type']).strip()
                base_m_type = raw_base_m_type
                
                # If the vehicle itself is natively carrying converted milk, collapse it to its target base type
                native_converted_type = None
                if " to " in raw_base_m_type.lower():
                    native_converted_type = raw_base_m_type
                    base_m_type = raw_base_m_type.split(" to ")[1].strip()
                    converted_types.add(native_converted_type)
                    
                total_qty = r.get('Dispatch Quantity', 0.0)
                if pd.notna(total_qty) and total_qty > 0:
                    key_base = (from_id, to_id, base_m_type)
                    dispatch_map_opt[key_base] = dispatch_map_opt.get(key_base, 0.0) + total_qty
                    
                    if native_converted_type:
                        orig_qty = r.get('Original Dispatch Quantity', total_qty)
                        if pd.notna(orig_qty) and orig_qty > 0:
                            key_conv_native = (from_id, to_id, base_m_type, native_converted_type)
                            converted_map_opt[key_conv_native] = converted_map_opt.get(key_conv_native, 0.0) + orig_qty
                
            bmc_dispatch_opt_rows = []
            sorted_conv_types = sorted(list(converted_types))
            for h in hubs:
                h_prods = h.get('products', [])
                if not h_prods and 'capacity' in h:
                    h_prods = [{'type': 'Cow Milk', 'capacity': h['capacity']}]
                
                hub_cap_map = {}
                hub_commodities_opt = set()
                for p_item in h_prods:
                    if p_item.get('type'):
                        hub_commodities_opt.add(p_item['type'])
                        hub_cap_map[p_item['type']] = float(p_item.get('capacity', 0.0))
                        
                for (from_id, to_id, prod_type) in dispatch_map_opt:
                    if from_id == h['id']:
                        hub_commodities_opt.add(prod_type)
                        
                for m in sorted(list(hub_commodities_opt)):
                    total_dispatch_opt = sum(dispatch_map_opt.get((h['id'], p['id'], m), 0.0) for p in plants)
                    total_cap = hub_cap_map.get(m, 0.0)
                    
                    if round(total_dispatch_opt, 2) > 0 or round(total_cap, 2) > 0:
                        row_dispatch_dict = {
                            'BMC': h.get('name', h['id']),
                            'Product': m,
                            'Total Capacity': round(total_cap, 2),
                            'Quantity': round(total_dispatch_opt, 2),
                            'Spare': round(total_cap - total_dispatch_opt, 2)
                        }
                        for p in plants:
                            p_name = p.get('name', p['id'])
                            disp_val = dispatch_map_opt.get((h['id'], p['id'], m), 0.0)
                            row_dispatch_dict[p_name] = round(disp_val, 2)
                            
                        # Add the converted milk columns
                        sum_conv_val = 0.0
                        for conv_type in sorted_conv_types:
                            conv_val = sum(converted_map_opt.get((h['id'], p['id'], m, conv_type), 0.0) for p in plants)
                            if conv_val > 0:
                                row_dispatch_dict[conv_type] = round(conv_val, 2)
                                sum_conv_val += conv_val
                            else:
                                row_dispatch_dict[conv_type] = 0.0
                                
                        row_dispatch_dict['Actual Quantity'] = round(total_dispatch_opt - sum_conv_val, 2)
                        bmc_dispatch_opt_rows.append(row_dispatch_dict)
                        
            df_bmc_wise_dispatch_opt = pd.DataFrame(bmc_dispatch_opt_rows)
            if not df_bmc_wise_dispatch_opt.empty:
                cols = ['BMC', 'Product', 'Total Capacity'] + sorted_conv_types + ['Actual Quantity', 'Quantity', 'Spare'] + [p_name for p_name in plant_names_list if p_name in df_bmc_wise_dispatch_opt.columns]
                df_bmc_wise_dispatch_opt = df_bmc_wise_dispatch_opt.reindex(columns=cols)
            else:
                df_bmc_wise_dispatch_opt = pd.DataFrame(columns=['BMC', 'Product', 'Total Capacity'] + sorted_conv_types + ['Actual Quantity', 'Quantity', 'Spare'] + plant_names_list)
                
                
            # --- Generate Plant Wise Dispatch (Max Utilized) ---
            plant_dispatch_opt_rows = []
            for p in plants:
                plant_commodities_opt = set()
                for (from_id, to_id, prod_type) in dispatch_map_opt:
                    if to_id == p['id']:
                        plant_commodities_opt.add(prod_type)
                        
                p_commodities = {}
                p_demands = p.get('demands', [])
                if p_demands:
                    for d in p_demands:
                        p_commodities[d['type']] = d.get('demand', 0.0)
                else:
                    inflow_milks = p.get('inflow_milks', [])
                    if not inflow_milks:
                        inflow_milks = [
                            {'type': 'Cow Milk', 'capacity': p.get('capacity', 10000)},
                            {'type': 'Buffalo Milk', 'capacity': p.get('capacity', 10000)}
                        ]
                    for m in inflow_milks:
                        if 'type' in m:
                            p_commodities[m['type']] = m.get('capacity', 0.0)
                            
                for m in list(plant_commodities_opt):
                    if m not in p_commodities:
                        p_commodities[m] = 0.0
                        
                for m, req_qty in sorted(p_commodities.items()):
                    total_dispatch_opt = sum(dispatch_map_opt.get((h['id'], p['id'], m), 0.0) for h in hubs)
                    
                    if req_qty == 0:
                        dispatch_pct = ""
                    else:
                        dispatch_pct = round((total_dispatch_opt / req_qty * 100.0), 2)
                        
                    row_dispatch_dict = {
                        'Plant': p.get('name', p['id']),
                        'Product': m,
                    }
                    
                    # Add converted milk columns (aggregated across all hubs sending to this plant)
                    sum_conv_val = 0.0
                    for conv_type in sorted_conv_types:
                        conv_val = sum(converted_map_opt.get((h['id'], p['id'], m, conv_type), 0.0) for h in hubs)
                        if conv_val > 0:
                            row_dispatch_dict[conv_type] = round(conv_val, 2)
                            sum_conv_val += conv_val
                        else:
                            row_dispatch_dict[conv_type] = 0.0
                            
                    row_dispatch_dict['Required Quantity'] = round(req_qty, 2)
                    row_dispatch_dict['Actual Quantity'] = round(total_dispatch_opt - sum_conv_val, 2)
                    row_dispatch_dict['Fullfilled Quantity'] = round(total_dispatch_opt, 2)
                    row_dispatch_dict['Fullfilled Percentage'] = dispatch_pct
                    
                    for h in hubs:
                        h_name = h.get('name', h['id'])
                        disp_val = dispatch_map_opt.get((h['id'], p['id'], m), 0.0)
                        row_dispatch_dict[h_name] = round(disp_val, 2)
                        
                    plant_dispatch_opt_rows.append(row_dispatch_dict)
                    
            df_plant_wise_dispatch_opt = pd.DataFrame(plant_dispatch_opt_rows)
            if not df_plant_wise_dispatch_opt.empty:
                cols = ['Plant', 'Product', 'Required Quantity'] + sorted_conv_types + ['Actual Quantity', 'Fullfilled Quantity', 'Fullfilled Percentage'] + [h_name for h_name in bmc_names_list if h_name in df_plant_wise_dispatch_opt.columns]
                df_plant_wise_dispatch_opt = df_plant_wise_dispatch_opt.reindex(columns=cols)
            else:
                df_plant_wise_dispatch_opt = pd.DataFrame(columns=['Plant', 'Product', 'Required Quantity'] + sorted_conv_types + ['Actual Quantity', 'Fullfilled Quantity', 'Fullfilled Percentage'] + bmc_names_list)

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_vehicle_wise_report.to_excel(writer, sheet_name='Vehicle Wise Report', index=False)
                df_kpi.to_excel(writer, sheet_name='KPI Summary', index=False)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                df_nodes.to_excel(writer, sheet_name='Nodes', index=False)
                df_routes.to_excel(writer, sheet_name='Routes', index=False)
                df_routes_optimized.to_excel(writer, sheet_name='Routes (Max Utilized)', index=False)
                df_plant_report.to_excel(writer, sheet_name='Plant Consumption Report', index=False)
                df_plant_received_report.to_excel(writer, sheet_name='Plant Received Report', index=False)
                df_hub_report.to_excel(writer, sheet_name='BMC Supply Report', index=False)
                df_hub_dispatch_report.to_excel(writer, sheet_name='BMC Dispatch Report', index=False)
                df_hub_to_plant.to_excel(writer, sheet_name='Hub To Plant', index=False)
                df_total_supply.to_excel(writer, sheet_name='Total Supply', index=False)
                df_vehicle_bifurcation.to_excel(writer, sheet_name='Vehicle Wise Bifurcation', index=False)
                df_bmc_wise_alloc.to_excel(writer, sheet_name='BMC Wise Allocation', index=False)
                df_bmc_wise_dispatch.to_excel(writer, sheet_name='BMC Wise Dispatch', index=False)
                df_bmc_wise_dispatch_opt.to_excel(writer, sheet_name='BMC Wise Dispatch (Max Util)', index=False)
                df_plant_wise_alloc.to_excel(writer, sheet_name='Plant Wise Allocation', index=False)
                df_plant_wise_dispatch.to_excel(writer, sheet_name='Plant Wise Dispatch', index=False)
                df_plant_wise_dispatch_opt.to_excel(writer, sheet_name='Plant Wise Dispatch (Max Util)', index=False)
                df_veh_alloc.to_excel(writer, sheet_name='BMC Vehicle Allocation', index=False)
                
                # --- New Supplier Report ---
                dispatch_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'cases', 'Summarized_Dispatch_Details_Supplier-sup-plant.xlsx')
                df_dispatch = pd.DataFrame()
                if os.path.exists(dispatch_file_path):
                    try:
                        df_dispatch = pd.read_excel(dispatch_file_path)
                        # Ensure columns are strings for matching
                        if 'Supplier' in df_dispatch.columns:
                            df_dispatch['Supplier'] = df_dispatch['Supplier'].astype(str).str.replace(r'\.0$', '', regex=True)
                        if 'Supplier Code' in df_dispatch.columns:
                            df_dispatch['Supplier Code'] = df_dispatch['Supplier Code'].astype(str).str.replace(r'\.0$', '', regex=True)
                    except Exception as e:
                        print("Error reading Summarized_Dispatch_Details_Supplier:", e)

                sup_code_to_name = {}
                if excel_file_path and os.path.exists(excel_file_path):
                    try:
                        df_dist = pd.read_excel(excel_file_path, sheet_name='Distance')
                        for _, row in df_dist.iterrows():
                            sc = str(row.get('Supplier Code', '')).strip()
                            if sc and sc not in sup_code_to_name:
                                sup_code_to_name[sc] = str(row.get('Supplier', ''))
                    except Exception as e:
                        print("Error reading Distance sheet for Supplier Report:", e)

                supplier_report_rows = []
                
                # Assign Supplier Code to df_total_supply
                df_total_supply_sup = df_total_supply.copy()
                df_total_supply_sup['Supplier Code'] = df_total_supply_sup['From Node ID'].apply(
                    lambda x: str(x).split('_')[0].strip() if len(str(x).split('_')) >= 2 else str(x).strip()
                )

                for sup_code, group in df_total_supply_sup.groupby('Supplier Code'):
                    total_tankers = group['Total Vehicles'].sum()
                    
                    if 'Total Distance' in group.columns:
                        total_distance = group['Total Distance'].sum()
                    elif 'Dispatch Quantity' in group.columns:
                        total_distance = group.loc[group['Dispatch Quantity'] > 0, 'Distance (km)'].sum()
                    else:
                        total_distance = group['Distance (km)'].sum()

                    if 'Dispatch Quantity' in group.columns:
                        total_supply = group['Dispatch Quantity'].sum()
                        left_quantity = (group['Flow'] - group['Dispatch Quantity']).sum()
                    else:
                        total_supply = group['Flow'].sum()
                        left_quantity = group['Flow'].sum()
                    
                    veh_counts = {}
                    for vc in veh_cols:
                        if vc in group.columns:
                            veh_counts[vc] = int(group[vc].sum())
                        else:
                            veh_counts[vc] = 0
                            
                    sup_name = sup_code_to_name.get(sup_code, '')

                    row_dict = {
                        'Supplier': sup_name,
                        'Supplier Code': sup_code,
                    }
                    
                    # Add individual vehicle columns
                    for vc in veh_cols:
                        row_dict[vc] = veh_counts.get(vc, 0)
                        
                    row_dict['Total Vehicles'] = total_tankers
                    row_dict['Total Distance'] = total_distance
                    row_dict['Total Supply'] = total_supply
                    row_dict['Left Quantity'] = left_quantity

                    # Join with df_dispatch
                    if not df_dispatch.empty and sup_code:
                        sup_col = 'Supplier Code' if 'Supplier Code' in df_dispatch.columns else 'Supplier'
                        if sup_col in df_dispatch.columns:
                            match = df_dispatch[df_dispatch[sup_col] == sup_code]
                            if not match.empty:
                                match_row = match.iloc[0].to_dict()
                                for k, v in match_row.items():
                                    if k not in row_dict:
                                        row_dict[k] = v
                                    else:
                                        row_dict[f'Dispatch_{k}'] = v
                    
                    supplier_report_rows.append(row_dict)
                
                df_supplier_summary = pd.DataFrame(supplier_report_rows)
                df_supplier_summary.to_excel(writer, sheet_name='Supplier Report', index=False)

                # --- BMC Vehicle Allocation (Max Utilized) ---
                veh_cols_opt = [f'{vc} Vehicles' for vc in vehicle_limits_map.get('global_caps', {}).keys()]
                cols_to_extract_opt = [
                    'From Node ID', 'From Name', 'To Node ID', 'To Name', 
                    'Product / Milk Type', 'Flow', 'Original Dispatch Quantity', 'Added Flow', 'Dispatch Quantity', 'Unit', 'Distance (km)', 'Transport Cost (₹)'
                ] + veh_cols_opt + [
                    'Total Vehicles', 'Total Vehicle Capacity (L)', 'Excess Vehicle Capacity (L)',
                    'SupplierCluster', 'SupplierSubCluster', 'Strategy',
                    'FlowLowMarginPercentage', 'FlowHighMarginPercentage',
                    'MinimumFlowQuantity', 'MaximumFlowQuantity', 'VehicleReason'
                ]
                valid_cols_opt = [c for c in cols_to_extract_opt if c in df_opt_active.columns]
                
                added_flow_cols = []
                i = 1
                while f'Added Flow {i}' in df_opt_active.columns:
                    added_flow_cols.extend([f'Added Flow {i}', f'Added Flow Source {i}', f'Added Flow Milk Type {i}'])
                    i += 1
                    
                df_veh_alloc_opt = df_opt_active[valid_cols_opt + added_flow_cols].copy()
                if 'Flow' in df_veh_alloc_opt.columns and 'Dispatch Quantity' in df_veh_alloc_opt.columns:
                    df_veh_alloc_opt['Left Quantity'] = df_veh_alloc_opt['Flow'] - df_veh_alloc_opt['Dispatch Quantity']
                
                df_veh_alloc_opt = df_veh_alloc_opt.rename(columns={
                    'From Node ID': 'BMC ID',
                    'From Name': 'BMC Name',
                    'To Node ID': 'Plant ID',
                    'To Name': 'Plant Name',
                    'Flow': 'Flow Quantity',
                    'Original Dispatch Quantity': 'Original Flow'
                })
                df_veh_alloc_opt.to_excel(writer, sheet_name='BMC Vehicle Allocation (Max Util)', index=False)

                # --- Total Supply Report (Max Utilized) ---
                df_total_supply_opt_temp = df_opt_active.copy()
                if not df_total_supply_opt_temp.empty:
                    df_total_supply_opt_temp['Base Milk'] = df_total_supply_opt_temp['Product / Milk Type'].apply(lambda x: str(x).split(' to ')[1].strip() if ' to ' in str(x).lower() else str(x).strip())
                    
                    cols_to_drop = ['Product / Milk Type', 'Route ID', 'Target Milk']
                    i = 1
                    while f'Added Flow {i}' in df_total_supply_opt_temp.columns:
                        cols_to_drop.extend([f'Added Flow {i}', f'Added Flow Source {i}', f'Added Flow Milk Type {i}', f'Added Flow Original Milk Type {i}'])
                        i += 1
                    df_total_supply_opt_temp = df_total_supply_opt_temp.drop(columns=cols_to_drop, errors='ignore')
                    
                    group_cols = ['From Node ID', 'To Node ID', 'Base Milk']
                    agg_dict = {}
                    for col in df_total_supply_opt_temp.columns:
                        if col in group_cols:
                            continue
                        if col in [
                            'Flow', 'Dispatch Quantity', 
                            'Original Flow', 'Added Flow', 
                            'Original Dispatch Quantity', 'Added Dispatch Quantity', 
                            'Original Excess Vehicle Capacity (L)', 'Excess Vehicle Capacity (L)', 
                            'Transport Cost (₹)', 'Total Vehicles', 'Total Vehicle Capacity (L)'
                        ] or str(col).endswith(' Vehicles'):
                            agg_dict[col] = 'sum'
                        else:
                            agg_dict[col] = 'first'
                    
                    df_total_supply_opt = df_total_supply_opt_temp.groupby(group_cols, observed=True).agg(agg_dict).reset_index()
                    df_total_supply_opt['Route ID'] = 'route_' + df_total_supply_opt['From Node ID'].astype(str) + '_' + df_total_supply_opt['To Node ID'].astype(str) + '_' + df_total_supply_opt['Base Milk'].astype(str).str.replace(' ', '_')
                    
                    original_cols = [c for c in df_opt_active.columns if c not in ['Target Milk'] and not str(c).startswith('Added Flow ')]
                    if 'Product / Milk Type' in original_cols:
                        idx = original_cols.index('Product / Milk Type')
                        new_cols = original_cols[:idx] + ['Base Milk'] + original_cols[idx+1:]
                    else:
                        new_cols = original_cols + ['Base Milk']
                        
                    df_total_supply_opt = df_total_supply_opt[[c for c in new_cols if c in df_total_supply_opt.columns]]
                    
                    if 'Dispatch Quantity' in df_total_supply_opt.columns:
                        df_total_supply_opt['Total Distance'] = df_total_supply_opt.apply(lambda r: (r.get('Total Vehicles', 0) * r.get('Distance (km)', 0)) if r.get('Dispatch Quantity', 0) > 0 else 0.0, axis=1)
                    else:
                        df_total_supply_opt['Total Distance'] = df_total_supply_opt.get('Total Vehicles', 0) * df_total_supply_opt.get('Distance (km)', 0)
                        
                    global_caps = vehicle_limits_map.get('global_caps', {})
                    for vc, cap in global_caps.items():
                        v_col = f'{vc} Vehicles'
                        new_col = f'{vc} Dispatch Capacity'
                        if v_col in df_total_supply_opt.columns:
                            tot_cap_s = df_total_supply_opt['Total Vehicle Capacity (L)'].replace({0: 1})
                            mask_cap_pos = df_total_supply_opt['Total Vehicle Capacity (L)'] > 0
                            
                            if 'Dispatch Quantity' in df_total_supply_opt.columns:
                                df_total_supply_opt[new_col] = round((df_total_supply_opt['Dispatch Quantity'] * (df_total_supply_opt[v_col] * cap) / tot_cap_s).where(mask_cap_pos, 0.0), 2)
                                if 'Original Dispatch Quantity' in df_total_supply_opt.columns:
                                    df_total_supply_opt[f'{vc} Original Dispatch Capacity'] = round((df_total_supply_opt['Original Dispatch Quantity'] * (df_total_supply_opt[v_col] * cap) / tot_cap_s).where(mask_cap_pos, 0.0), 2)
                                    df_total_supply_opt[f'{vc} Added Dispatch Capacity'] = round((df_total_supply_opt['Added Dispatch Quantity'] * (df_total_supply_opt[v_col] * cap) / tot_cap_s).where(mask_cap_pos, 0.0), 2)
                            else:
                                df_total_supply_opt[new_col] = round((df_total_supply_opt['Flow'] * (df_total_supply_opt[v_col] * cap) / tot_cap_s).where(mask_cap_pos, 0.0), 2)
                else:
                    cols = [c if c != 'Product / Milk Type' else 'Base Milk' for c in df_opt_active.columns if c != 'Target Milk' and not str(c).startswith('Added Flow ')] + ['Total Distance']
                    global_caps = vehicle_limits_map.get('global_caps', {})
                    for vc in global_caps.keys():
                        cols.extend([f'{vc} Dispatch Capacity', f'{vc} Original Dispatch Capacity', f'{vc} Added Dispatch Capacity'])
                    df_total_supply_opt = pd.DataFrame(columns=cols)

                df_total_supply_opt.to_excel(writer, sheet_name='Total Supply (Max Util)', index=False)
                    
                # --- Vehicle Wise Bifurcation (Max Utilized) ---
                veh_bifurcation_opt_rows = []
                global_caps = vehicle_limits_map.get('global_caps', {})
                for vc, cap in global_caps.items():
                    v_col = f'{vc} Vehicles'
                    if v_col in df_total_supply_opt.columns and df_total_supply_opt[v_col].sum() > 0:
                        total_veh = df_total_supply_opt[v_col].sum()
                        total_dist = (df_total_supply_opt['Distance (km)'] * df_total_supply_opt[v_col]).sum()
                        
                        disp_cap_col = f'{vc} Dispatch Capacity'
                        orig_disp_cap_col = f'{vc} Original Dispatch Capacity'
                        add_disp_cap_col = f'{vc} Added Dispatch Capacity'
                        
                        if disp_cap_col in df_total_supply_opt.columns:
                            total_supply_v = df_total_supply_opt[disp_cap_col].sum()
                        else:
                            supply_series = df_total_supply_opt['Flow'] * (df_total_supply_opt[v_col] * cap) / df_total_supply_opt['Total Vehicle Capacity (L)'].replace({0: 1})
                            total_supply_v = supply_series.where(df_total_supply_opt['Total Vehicle Capacity (L)'] > 0, 0).sum()
                            
                        total_orig_supply_v = df_total_supply_opt[orig_disp_cap_col].sum() if orig_disp_cap_col in df_total_supply_opt.columns else 0.0
                        total_add_supply_v = df_total_supply_opt[add_disp_cap_col].sum() if add_disp_cap_col in df_total_supply_opt.columns else 0.0
                        
                        veh_bifurcation_opt_rows.append({
                            'Vehicle Category': vc,
                            'Total Distance': total_dist,
                            'Original Dispatch Capacity': total_orig_supply_v,
                            'Added Dispatch Capacity': total_add_supply_v,
                            'Dispatch Capacity': total_supply_v,
                            'TotalVehicle': total_veh
                        })
                        
                df_vehicle_bifurcation_opt = pd.DataFrame(veh_bifurcation_opt_rows)
                if not df_vehicle_bifurcation_opt.empty:
                    total_row_bif = {
                        'Vehicle Category': 'Grand Total',
                        'Total Distance': df_vehicle_bifurcation_opt['Total Distance'].sum(),
                        'Original Dispatch Capacity': df_vehicle_bifurcation_opt['Original Dispatch Capacity'].sum(),
                        'Added Dispatch Capacity': df_vehicle_bifurcation_opt['Added Dispatch Capacity'].sum(),
                        'Dispatch Capacity': df_vehicle_bifurcation_opt['Dispatch Capacity'].sum(),
                        'TotalVehicle': df_vehicle_bifurcation_opt['TotalVehicle'].sum()
                    }
                    df_vehicle_bifurcation_opt = pd.concat([df_vehicle_bifurcation_opt, pd.DataFrame([total_row_bif])], ignore_index=True)
                else:
                    df_vehicle_bifurcation_opt = pd.DataFrame(columns=['Vehicle Category', 'Total Distance', 'Original Dispatch Capacity', 'Added Dispatch Capacity', 'Dispatch Capacity', 'TotalVehicle'])
                    
                df_vehicle_bifurcation_opt.to_excel(writer, sheet_name='Vehicle Wise Bifurcation (Max Util)', index=False)

                # --- Supplier Report (Max Utilized) ---
                supplier_report_opt_rows = []
                if not df_total_supply_opt.empty:
                    df_total_supply_opt_sup = df_total_supply_opt.copy()
                    df_total_supply_opt_sup['Supplier Code'] = df_total_supply_opt_sup['From Node ID'].apply(
                        lambda x: str(x).split('_')[0].strip() if len(str(x).split('_')) >= 2 else str(x).strip()
                    )
                    
                    if 'Detailed Reason' in df_total_supply_opt_sup.columns:
                        df_total_supply_opt_sup = df_total_supply_opt_sup.drop(columns=['Detailed Reason'])

                    for sup_code, group in df_total_supply_opt_sup.groupby('Supplier Code'):
                        total_tankers = group['Total Vehicles'].sum()
                        
                        if 'Total Distance' in group.columns:
                            total_distance = group['Total Distance'].sum()
                        elif 'Dispatch Quantity' in group.columns:
                            total_distance = group.loc[group['Dispatch Quantity'] > 0, 'Distance (km)'].sum()
                        else:
                            total_distance = group['Distance (km)'].sum()
                            
                        if 'Dispatch Quantity' in group.columns:
                            total_supply = group['Dispatch Quantity'].sum()
                            left_quantity = (group['Flow'] - group['Dispatch Quantity']).sum()
                        else:
                            total_supply = group['Flow'].sum()
                            left_quantity = group['Flow'].sum()
                            
                        veh_counts = {}
                        veh_cols = [f"{vc} Vehicles" for vc in global_caps.keys()]
                        for vc, vc_col in zip(global_caps.keys(), veh_cols):
                            if vc_col in group.columns:
                                veh_counts[vc] = int(group[vc_col].sum())
                            else:
                                veh_counts[vc] = 0
                                
                        sup_name = sup_code_to_name.get(sup_code, '')
                        
                        row_dict = {
                            'Supplier Code': sup_code,
                            'Supplier': sup_name,
                        }
                        
                        for vc in veh_cols:
                            row_dict[vc] = veh_counts.get(vc.replace(' Vehicles', ''), 0)
                            
                        row_dict['Total Vehicles'] = total_tankers
                        row_dict['Total Distance'] = total_distance
                        row_dict['Total Supply'] = total_supply
                        row_dict['Left Quantity'] = left_quantity
                        
                        if 'df_dispatch' in locals() and not df_dispatch.empty and sup_code:
                            sup_col = 'Supplier Code' if 'Supplier Code' in df_dispatch.columns else 'Supplier'
                            if sup_col in df_dispatch.columns:
                                match = df_dispatch[df_dispatch[sup_col] == sup_code]
                                if not match.empty:
                                    match_row = match.iloc[0].to_dict()
                                    for k, v in match_row.items():
                                        if k not in ['Supplier', 'Supplier Code']:
                                            if k not in row_dict:
                                                row_dict[k] = v
                                            else:
                                                row_dict[f'Dispatch_{k}'] = v
                                                
                        supplier_report_opt_rows.append(row_dict)
                        
                df_supplier_summary_opt = pd.DataFrame(supplier_report_opt_rows)
                df_supplier_summary_opt.to_excel(writer, sheet_name='Supplier Report (Max Util)', index=False)
                
                try:
                    from openpyxl.styles import PatternFill
                    from collections import Counter
                    base_route_ids = [str(r).split('_extra_')[0] if '_extra_' in str(r) else str(r) for r in df_routes_optimized['Route ID']]
                    counts = Counter(base_route_ids)
                    shared_base_ids = set(k for k, v in counts.items() if v > 1)
                    
                    highlight_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    worksheet = writer.sheets['Routes (Max Utilized)']
                    for row_idx, base_id in enumerate(base_route_ids):
                        if base_id in shared_base_ids:
                            excel_row = row_idx + 2
                            for col_idx in range(1, len(df_routes_optimized.columns) + 1):
                                worksheet.cell(row=excel_row, column=col_idx).fill = highlight_fill
                except Exception as e:
                    print("Error highlighting routes:", str(e))
                
            #update_job_completed(job_id, output_filename, res['summary'])
        #else:
            #update_job_failed(job_id, f"Solver solved to infeasible status: {res.get('status')}")
            
    except Exception as e:
        import traceback
        error_msg = f"Exception: {str(e)}\n{traceback.format_exc()}"
        print("Job processing failed:", error_msg)
        #update_job_failed(job_id, error_msg[:1000])







